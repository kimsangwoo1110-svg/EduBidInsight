"""Excel and PDF exports for school analytics dashboards."""

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


class DashboardExportService:
    """Create portable dashboard snapshots without accessing the GUI."""

    @classmethod
    def export_excel(cls, file_path, school_name, summary):
        workbook = Workbook()
        overview = workbook.active
        overview.title = "요약"
        overview.append(["학교 분석 대시보드", school_name])
        overview.append(["항목", "값"])
        for label, value in cls._summary_rows(summary):
            overview.append([label, value])

        cls._append_statistics_sheet(
            workbook,
            "품목 통계",
            ["품목", "계약수", "수량", "금액", "비중(%)"],
            [
                [row["product"], row["count"], row["quantity"], row["amount"], row["share"]]
                for row in summary["product_statistics"]
            ],
        )
        cls._append_statistics_sheet(
            workbook,
            "업체 통계",
            ["업체", "계약수", "수량", "금액", "비중(%)"],
            [
                [row["vendor"], row["count"], row["quantity"], row["amount"], row["share"]]
                for row in summary["vendor_statistics"]
            ],
        )
        cls._append_statistics_sheet(
            workbook,
            "연도별 추이",
            ["연도", "프로젝트수", "프로젝트 예산", "계약수", "계약금액"],
            [
                [
                    row["year"],
                    row["project_count"],
                    row["project_budget"],
                    row["contract_count"],
                    row["contract_amount"],
                ]
                for row in summary["yearly_trend"]
            ],
        )
        cls._append_statistics_sheet(
            workbook,
            "영업 기회",
            ["대상", "추천", "점수", "우선순위", "근거"],
            [
                [
                    row["project_name"],
                    row["recommendation"],
                    row["score"],
                    row["priority"],
                    row["reason"],
                ]
                for row in summary["opportunity"]["insights"]
            ],
        )
        business = summary.get("business_insight")
        if business:
            cls._append_business_sheets(workbook, business)
        cls._style_sheet(overview)
        workbook.save(file_path)
        workbook.close()
        return file_path

    @classmethod
    def export_pdf(cls, file_path, school_name, summary):
        """Write a compact Korean PDF summary using the standard Korea1 CMap."""
        lines = [
            f"EduBid Insight 학교 분석 요약 - {school_name}",
            "",
            *[
                f"{label}: {cls._pdf_value(label, value)}"
                for label, value in cls._summary_rows(summary)
            ],
        ]
        business = summary.get("business_insight")
        if business:
            lines.extend(
                [
                    "",
                    f"Business Insight: {business['summary']}",
                    f"Explanation: {business['explanation']['text']}",
                    "추천 제품: "
                    + ", ".join(
                        recommendation["product"]
                        for recommendation in business["recommended_products"][:3]
                    ),
                    "주요 위험: "
                    + " / ".join(risk["risk"] for risk in business["risks"][:2]),
                    "다음 행동: "
                    + " / ".join(
                        action["action"] for action in business["next_actions"][:2]
                    ),
                ]
            )
        content_commands = ["BT", "/F1 15 Tf", "50 790 Td"]
        for index, line in enumerate(lines[:31]):
            if index:
                content_commands.append("0 -23 Td")
            encoded = line[:90].encode("utf-16-be").hex().upper()
            content_commands.append(f"<{encoded}> Tj")
        content_commands.append("ET")
        content = "\n".join(content_commands).encode("ascii")

        objects = [
            b"<< /Type /Catalog /Pages 2 0 R >>",
            b"<< /Type /Pages /Kids [3 0 R] /Count 1 >>",
            (
                b"<< /Type /Page /Parent 2 0 R /MediaBox [0 0 595 842] "
                b"/Resources << /Font << /F1 5 0 R >> >> /Contents 4 0 R >>"
            ),
            b"<< /Length " + str(len(content)).encode("ascii") + b" >>\nstream\n"
            + content
            + b"\nendstream",
            (
                b"<< /Type /Font /Subtype /Type0 /BaseFont /HYGoThic-Medium "
                b"/Encoding /UniKS-UCS2-H /DescendantFonts [6 0 R] >>"
            ),
            (
                b"<< /Type /Font /Subtype /CIDFontType0 /BaseFont /HYGoThic-Medium "
                b"/CIDSystemInfo << /Registry (Adobe) /Ordering (Korea1) "
                b"/Supplement 2 >> /DW 1000 >>"
            ),
        ]
        pdf = bytearray(b"%PDF-1.4\n%\xE2\xE3\xCF\xD3\n")
        offsets = [0]
        for object_number, body in enumerate(objects, start=1):
            offsets.append(len(pdf))
            pdf.extend(f"{object_number} 0 obj\n".encode("ascii"))
            pdf.extend(body)
            pdf.extend(b"\nendobj\n")
        xref_offset = len(pdf)
        pdf.extend(f"xref\n0 {len(objects) + 1}\n".encode("ascii"))
        pdf.extend(b"0000000000 65535 f \n")
        for offset in offsets[1:]:
            pdf.extend(f"{offset:010d} 00000 n \n".encode("ascii"))
        pdf.extend(
            (
                f"trailer\n<< /Size {len(objects) + 1} /Root 1 0 R >>\n"
                f"startxref\n{xref_offset}\n%%EOF\n"
            ).encode("ascii")
        )
        with open(file_path, "wb") as pdf_file:
            pdf_file.write(pdf)
        return file_path

    @staticmethod
    def _summary_rows(summary):
        kpis = summary["kpis"]
        projects = summary["project_summary"]
        contracts = summary["contract_summary"]
        cycle = summary["purchase_cycle"]
        opportunity = summary["opportunity"]
        return [
            ("프로젝트 수", kpis["projects"]),
            ("프로젝트 예산", kpis["project_budget"]),
            ("진행중 프로젝트", projects["status_counts"]["진행중"]),
            ("예정 프로젝트", projects["status_counts"]["예정"]),
            ("완료 프로젝트", projects["status_counts"]["완료"]),
            ("보류 프로젝트", projects["status_counts"]["보류"]),
            ("계약 수", kpis["contracts"]),
            ("계약 총액", kpis["contract_amount"]),
            ("평균 계약금액", contracts["average_amount"]),
            ("최대 계약금액", contracts["maximum_amount"]),
            ("거래 업체 수", kpis["vendors"]),
            ("평균 구매 주기", cycle["average_days"]),
            ("최근 구매일", cycle["last_purchase_date"] or "-"),
            ("다음 예상 구매일", cycle["next_expected_date"] or "-"),
            ("Opportunity Score", opportunity["score"]),
            ("영업 기회 수", opportunity["opportunity_count"]),
            ("높은 우선순위", opportunity["high_priority_count"]),
        ]

    @staticmethod
    def _pdf_value(label, value):
        if value is None:
            return "-"
        if "예산" in label or "금액" in label or "총액" in label:
            return f"{float(value):,.0f}원"
        if label == "평균 구매 주기":
            return f"{float(value):,.1f}일"
        return str(value)

    @classmethod
    def _append_statistics_sheet(cls, workbook, title, headers, rows):
        worksheet = workbook.create_sheet(title)
        worksheet.append(headers)
        for row in rows:
            worksheet.append(row)
        cls._style_sheet(worksheet)

    @classmethod
    def _append_business_sheets(cls, workbook, business):
        overview = workbook.create_sheet("BI 요약")
        overview.append(["Business Insight", business["summary"]])
        overview.append(["Explanation", business["explanation"]["text"]])
        overview.append(["Score", business["score"]])
        overview.append(["생성일", business["created_at"]])
        cls._style_sheet(overview)

        cls._append_statistics_sheet(
            workbook,
            "BI 추천",
            ["제품", "신뢰도", "우선순위", "근거"],
            [
                [row["product"], row["confidence"], row["priority"], row["reason"]]
                for row in business["recommended_products"]
            ],
        )
        cls._append_statistics_sheet(
            workbook,
            "BI 위험과 행동",
            ["구분", "수준/시점", "내용", "대응/근거"],
            [
                ["위험", row["level"], row["risk"], row["mitigation"]]
                for row in business["risks"]
            ]
            + [
                ["행동", row["timing"], row["action"], row["reason"]]
                for row in business["next_actions"]
            ],
        )
        cls._append_statistics_sheet(
            workbook,
            "기회 타임라인",
            ["날짜", "유형", "기회", "상세"],
            [
                [row["date"], row["type"], row["title"], row["detail"]]
                for row in business["opportunity_timeline"]
            ],
        )
        cls._append_statistics_sheet(
            workbook,
            "우선순위 매트릭스",
            ["항목", "영향도", "긴급도", "영역", "근거"],
            [
                [
                    row["item"],
                    row["impact"],
                    row["urgency"],
                    row["quadrant"],
                    row["reason"],
                ]
                for row in business["priority_matrix"]
            ],
        )

    @staticmethod
    def _style_sheet(worksheet):
        header_row = 2 if worksheet.title == "요약" else 1
        fill = PatternFill(fill_type="solid", fgColor="1F6AA5")
        for cell in worksheet[header_row]:
            cell.fill = fill
            cell.font = Font(bold=True, color="FFFFFF")
            cell.alignment = Alignment(horizontal="center")
        worksheet.freeze_panes = f"A{header_row + 1}"
        for column_cells in worksheet.columns:
            width = min(
                max((len(str(cell.value or "")) for cell in column_cells), default=0) + 3,
                45,
            )
            worksheet.column_dimensions[
                get_column_letter(column_cells[0].column)
            ].width = width
