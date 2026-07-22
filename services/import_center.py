"""Import Center profiles, workbook analysis, templates, and extended audit data.

This module is an integration boundary. Existing domain services remain the only
place that persists schools, contracts, projects, and CRM activities.
"""

from __future__ import annotations

import csv
import io
import json
import os
import tempfile
import threading
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from time import perf_counter

from openpyxl import Workbook, load_workbook

from core.app_settings import get_app_settings
from core.resources import resource_path
from services.base_import import BaseImport, ImportCancelled
from services.connectors.contract_import import ContractImportConnector
from services.connectors.education_office_import import (
    EDUCATION_OFFICE_COLUMNS,
    EDUCATION_OFFICE_FIELD_LABELS,
    EDUCATION_OFFICE_REQUIRED_FIELDS,
    EducationOfficeImport,
)
from services.connectors.g2b_import import (
    G2B_COLUMNS,
    G2B_FIELD_LABELS,
    G2B_REQUIRED_FIELDS,
    G2BImport,
)
from services.connectors.schoolmarket_import import (
    SCHOOLMARKET_COLUMNS,
    SCHOOLMARKET_FIELD_LABELS,
    SCHOOLMARKET_REQUIRED_FIELDS,
    SchoolMarketImport,
)
from services.import_history_service import ImportHistoryService
from services.import_school import import_school_data, integer_value
from services.sales_activity_service import SalesActivityService


SCHOOL_COLUMNS = (
    "school_code", "school_name", "school_type", "office", "region", "address",
    "homepage", "ai_school", "digital_school", "space_innovation", "green_smart",
    "student_count", "class_count",
)
SCHOOL_REQUIRED_FIELDS = ("school_code", "school_name")
SCHOOL_FIELD_LABELS = {
    "school_code": "학교코드", "school_name": "학교명", "school_type": "학교급",
    "office": "교육청", "region": "지역", "address": "주소", "homepage": "홈페이지",
    "ai_school": "AI중점", "digital_school": "디지털학교", "space_innovation": "공간혁신",
    "green_smart": "그린스마트", "student_count": "학생수", "class_count": "학급수",
}
SCHOOL_ALIASES = {
    "school_code": ("학교코드", "표준학교코드", "기관코드", "school code", "schoolcode"),
    "school_name": ("학교명", "학교 이름", "학교이름", "기관명", "school name", "schoolname"),
    "school_type": ("학교급", "학교유형", "학교종류", "school type"),
    "office": ("교육청", "교육청명", "시도교육청", "관할교육청", "office"),
    "region": ("지역", "지역명", "시도", "시군구", "region"),
    "address": ("주소", "소재지", "도로명주소", "address"),
    "homepage": ("홈페이지", "홈페이지주소", "웹사이트", "website", "homepage"),
    "ai_school": ("AI중점", "AI학교", "ai school"),
    "digital_school": ("디지털학교", "디지털선도학교", "digital school"),
    "space_innovation": ("공간혁신", "학교공간혁신", "space innovation"),
    "green_smart": ("그린스마트", "그린스마트학교", "green smart"),
    "student_count": ("학생수", "학생 수", "재학생수", "students"),
    "class_count": ("학급수", "학급 수", "classes"),
}

CRM_COLUMNS = (
    "school_code", "activity_date", "activity_type", "contact_person", "memo",
    "next_action_date", "status",
)
CRM_REQUIRED_FIELDS = ("school_code", "activity_date", "activity_type")
CRM_FIELD_LABELS = {
    "school_code": "학교코드", "activity_date": "활동일", "activity_type": "활동유형",
    "contact_person": "담당자", "memo": "메모", "next_action_date": "다음액션일",
    "status": "영업단계",
}
CRM_ALIASES = {
    "school_code": ("학교코드", "표준학교코드", "기관코드", "school code", "schoolcode"),
    "activity_date": ("활동일", "활동일자", "영업일", "일자", "activity date", "date"),
    "activity_type": ("활동유형", "활동 유형", "액션유형", "영업활동", "activity type"),
    "contact_person": ("담당자", "담당자명", "연락담당자", "contact", "contact person"),
    "memo": ("메모", "내용", "상담내용", "비고", "memo", "note"),
    "next_action_date": ("다음액션일", "후속일", "예정일", "next action", "follow up"),
    "status": ("영업단계", "단계", "상태", "pipeline", "status"),
}


@dataclass(frozen=True)
class ImportProfile:
    key: str
    title: str
    source: str
    template_name: str
    fields: tuple[str, ...]
    labels: dict[str, str]
    required: tuple[str, ...]
    adapter: type


class MappedWorkbookImport(BaseImport):
    """Base adapter for new imports that delegate writes to existing services."""

    source = ""
    fields = ()
    labels = {}
    aliases = {}
    required = ()

    def __init__(self, filename, sheet_name=None, mapping=None, history_service=None, cancel_event=None):
        super().__init__(self.source, filename, history_service)
        self.extension = Path(self.filename).suffix.lower()
        if self.extension not in {".xlsx", ".csv"}:
            raise ValueError("Excel(.xlsx) 또는 CSV(.csv) 파일만 지원합니다.")
        if not os.path.isfile(self.filename):
            raise ValueError("선택한 파일을 찾을 수 없습니다.")
        sheets = self.sheet_names(self.filename)
        self.sheet_name = sheet_name or sheets[0]
        headers = self.headers(self.filename, self.sheet_name)
        self.mapping = mapping or self.auto_map(headers)
        self.validate_mapping(self.mapping, headers)
        self.cancel_event = cancel_event or threading.Event()
        self._loaded = None

    @classmethod
    def sheet_names(cls, file_path):
        if Path(file_path).suffix.lower() == ".csv":
            return ["CSV"]
        workbook = load_workbook(file_path, read_only=True, data_only=True)
        try:
            return list(workbook.sheetnames)
        finally:
            workbook.close()

    @classmethod
    def headers(cls, file_path, sheet_name=None):
        if Path(file_path).suffix.lower() == ".csv":
            return list(csv.DictReader(io.StringIO(ContractImportConnector._read_csv_content(file_path))).fieldnames or [])
        workbook = load_workbook(file_path, read_only=True, data_only=True)
        try:
            sheet = workbook[sheet_name or workbook.sheetnames[0]]
            return [str(value or "").strip() for value in next(sheet.iter_rows(values_only=True), ()) if str(value or "").strip()]
        finally:
            workbook.close()

    @classmethod
    def auto_map(cls, headers):
        normalized = {cls._normalize(value): value for value in headers}
        used = set()
        mapping = {}
        for field in cls.fields:
            mapping[field] = ""
            for alias in cls.aliases[field]:
                candidate = normalized.get(cls._normalize(alias), "")
                if candidate and candidate not in used:
                    mapping[field] = candidate
                    used.add(candidate)
                    break
        return mapping

    @classmethod
    def validate_mapping(cls, mapping, headers):
        missing = [field for field in cls.required if not mapping.get(field)]
        if missing:
            raise ValueError("필수 열 매핑 누락: " + ", ".join(cls.labels[field] for field in missing))
        selected = [value for value in mapping.values() if value]
        invalid = [value for value in selected if value not in set(headers)]
        if invalid:
            raise ValueError("파일에 없는 열: " + ", ".join(invalid))
        if len(selected) != len(set(selected)):
            raise ValueError("하나의 원본 열을 여러 필드에 매핑할 수 없습니다.")
        return True

    def load(self):
        if self._loaded is None:
            self._loaded = list(self._rows())
        return self._loaded

    def validate(self):
        result = []
        for row_number, source_row in enumerate(self.load(), start=2):
            values = {field: source_row.get(header, "") if header else "" for field, header in self.mapping.items()}
            try:
                normalized = self.validate_values(values)
                error = ""
            except (TypeError, ValueError) as validation_error:
                normalized = values
                error = str(validation_error)
            missing = [field for field in self.required if not str(values.get(field) or "").strip()]
            result.append({"row_number": row_number, "contract": normalized, "source": source_row, "error": error, "missing_fields": missing})
        return result

    def preview(self, limit=100):
        return self.validate()[:max(0, int(limit))]

    def cancel(self):
        self.cancel_event.set()

    def save(self):
        """Concrete subclasses persist through their existing service in run()."""
        raise NotImplementedError("use run() for audited imports")

    def _rows(self):
        if self.extension == ".csv":
            yield from csv.DictReader(io.StringIO(ContractImportConnector._read_csv_content(self.filename)))
            return
        workbook = load_workbook(self.filename, read_only=True, data_only=True)
        try:
            sheet = workbook[self.sheet_name]
            rows = sheet.iter_rows(values_only=True)
            headers = [str(value or "").strip() for value in next(rows, ())]
            for values in rows:
                yield {header: values[index] if index < len(values) else None for index, header in enumerate(headers) if header}
        finally:
            workbook.close()

    @staticmethod
    def _normalize(value):
        return "".join(character for character in str(value or "").casefold() if character.isalnum())

    def validate_values(self, values):
        raise NotImplementedError


class SchoolImport(MappedWorkbookImport):
    source = "학교 가져오기"
    fields, labels, aliases, required = SCHOOL_COLUMNS, SCHOOL_FIELD_LABELS, SCHOOL_ALIASES, SCHOOL_REQUIRED_FIELDS

    def validate_values(self, values):
        normalized = dict(values)
        for field in self.required:
            normalized[field] = str(values.get(field) or "").strip()
            if not normalized[field]:
                raise ValueError(f"필수 값 누락: {self.labels[field]}")
        for field in ("student_count", "class_count"):
            try:
                normalized[field] = integer_value(values.get(field))
            except (TypeError, ValueError) as error:
                raise ValueError(f"{self.labels[field]}은 숫자여야 합니다.") from error
        return normalized

    def run(self, progress_callback=None):
        started = datetime.now().astimezone(); timer = perf_counter()
        rows = self.validate(); valid = [row for row in rows if not row["error"]]
        status, imported, skipped, failed, warnings = "FAILED", 0, 0, len(rows) - len(valid), []
        try:
            if self.cancel_event.is_set(): raise ImportCancelled()
            if progress_callback: progress_callback(stage="Validating...", processed=0, total=len(rows), percentage=10)
            workbook = Workbook(); sheet = workbook.active
            sheet.append([self.labels[field] for field in self.fields])
            for row in valid: sheet.append([row["contract"].get(field, "") for field in self.fields])
            with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as temporary:
                temporary_path = temporary.name
            try:
                workbook.save(temporary_path); workbook.close()
                result = import_school_data(temporary_path)
            finally:
                try: os.unlink(temporary_path)
                except OSError: pass
            imported = result["inserted"] + result["updated"]
            skipped = result["skipped"]
            failed += result["errors"]
            status = "PARTIAL" if failed or skipped else "SUCCESS"
            warnings = [f"Row {row['row_number']}: {row['error']}" for row in rows if row["error"]]
            if progress_callback: progress_callback(stage="Saving...", processed=len(rows), total=len(rows), percentage=100)
        except ImportCancelled:
            status, imported, skipped, failed, warnings = "Cancelled", 0, 0, 0, ["사용자가 가져오기를 취소했습니다."]
        except Exception as error:
            warnings = [f"{type(error).__name__}: {error}"]
        elapsed = perf_counter() - timer
        self.log(status, imported)
        return {"status": status, "started_at": started.isoformat(timespec="seconds"), "elapsed": elapsed, "imported": imported, "skipped": skipped, "failed": failed, "warnings": warnings, "cancelled": status == "Cancelled"}


class CRMImport(MappedWorkbookImport):
    source = "CRM 가져오기"
    fields, labels, aliases, required = CRM_COLUMNS, CRM_FIELD_LABELS, CRM_ALIASES, CRM_REQUIRED_FIELDS

    def validate_values(self, values):
        candidate = dict(values)
        candidate["status"] = str(candidate.get("status") or "Lead").strip()
        return SalesActivityService._validate(**candidate)

    def run(self, progress_callback=None):
        started = datetime.now().astimezone(); timer = perf_counter()
        rows = self.validate(); imported = failed = 0; warnings = []
        status = "FAILED"
        try:
            for processed, row in enumerate(rows, start=1):
                if self.cancel_event.is_set(): raise ImportCancelled()
                if row["error"]:
                    failed += 1; warnings.append(f"Row {row['row_number']}: {row['error']}")
                else:
                    SalesActivityService.add_activity(**row["contract"]); imported += 1
                if progress_callback: progress_callback(stage="Importing...", processed=processed, total=len(rows), percentage=round(processed * 100 / len(rows)) if rows else 100)
            status = "PARTIAL" if failed else "SUCCESS"
        except ImportCancelled:
            status = "Cancelled"; warnings.append("사용자가 가져오기를 취소했습니다.")
        except Exception as error:
            status = "FAILED"; warnings.append(f"{type(error).__name__}: {error}")
        elapsed = perf_counter() - timer
        self.log(status, imported)
        return {"status": status, "started_at": started.isoformat(timespec="seconds"), "elapsed": elapsed, "imported": imported, "skipped": 0, "failed": failed, "warnings": warnings, "cancelled": status == "Cancelled"}


PROFILES = {
    "school": ImportProfile("school", "학교 가져오기\nSchool Import", "School", "School_Import_Template.xlsx", SCHOOL_COLUMNS, SCHOOL_FIELD_LABELS, SCHOOL_REQUIRED_FIELDS, SchoolImport),
    "education_office": ImportProfile("education_office", "교육청 가져오기\nEducation Office Import", "Education Office", "Education_Office_Import_Template.xlsx", EDUCATION_OFFICE_COLUMNS, EDUCATION_OFFICE_FIELD_LABELS, EDUCATION_OFFICE_REQUIRED_FIELDS, EducationOfficeImport),
    "schoolmarket": ImportProfile("schoolmarket", "학교장터 가져오기\nSchool Market Import", "School Market", "School_Market_Import_Template.xlsx", SCHOOLMARKET_COLUMNS, SCHOOLMARKET_FIELD_LABELS, SCHOOLMARKET_REQUIRED_FIELDS, SchoolMarketImport),
    "g2b": ImportProfile("g2b", "나라장터 가져오기\nNaraJangteo Import", "NaraJangteo", "NaraJangteo_Import_Template.xlsx", G2B_COLUMNS, G2B_FIELD_LABELS, G2B_REQUIRED_FIELDS, G2BImport),
    "crm": ImportProfile("crm", "CRM 가져오기\nCRM Import", "CRM", "CRM_Import_Template.xlsx", CRM_COLUMNS, CRM_FIELD_LABELS, CRM_REQUIRED_FIELDS, CRMImport),
}


def template_path(profile_key):
    return Path(resource_path(Path("templates") / PROFILES[profile_key].template_name))


def failed_rows(importer):
    """Return original worksheet rows that failed pre-import validation."""
    rows = []
    source_rows = list(importer.load())
    for item in importer.validate():
        if item.get("error"):
            index = max(0, int(item.get("row_number", 2)) - 2)
            original = item.get("source") or (source_rows[index] if index < len(source_rows) else {})
            rows.append({**dict(original), "검증 오류 / Validation Error": item["error"]})
    return rows


def export_failed_rows(rows, destination):
    if not rows:
        raise ValueError("내보낼 실패 행이 없습니다.")
    headers = list(dict.fromkeys(key for row in rows for key in row))
    workbook = Workbook(); sheet = workbook.active; sheet.title = "Failed Rows"
    sheet.append(headers)
    for row in rows: sheet.append([row.get(header, "") for header in headers])
    sheet.freeze_panes = "A2"; sheet.auto_filter.ref = sheet.dimensions
    workbook.save(destination); workbook.close()
    return destination


class ImportRunStore:
    """Persist richer Import Center audit fields without changing SQLite schema."""

    def __init__(self, path=None):
        settings = get_app_settings()
        self.path = Path(path or settings.settings_path.parent / "import_history.json")

    def record(self, profile, filename, analyzed_rows, summary):
        entry = {
            "date": summary.get("started_at") or datetime.now().astimezone().isoformat(timespec="seconds"),
            "source": profile.source,
            "filename": os.path.basename(filename),
            "rows": int(analyzed_rows or 0),
            "success": int(summary.get("imported", 0) or 0),
            "failed": int(summary.get("failed", 0) or 0),
            "duration": float(summary.get("elapsed", 0) or 0),
            "status": str(summary.get("status") or ""),
        }
        history = self.history(limit=None)
        history.insert(0, entry)
        self.path.parent.mkdir(parents=True, exist_ok=True)
        temporary = self.path.with_suffix(self.path.suffix + ".tmp")
        temporary.write_text(json.dumps(history[:500], ensure_ascii=False, indent=2), encoding="utf-8")
        os.replace(temporary, self.path)
        return entry

    def history(self, limit=100):
        try:
            loaded = json.loads(self.path.read_text(encoding="utf-8"))
        except (FileNotFoundError, OSError, json.JSONDecodeError):
            loaded = []
        rows = [dict(item) for item in loaded if isinstance(item, dict)]
        return rows if limit is None else rows[:max(0, int(limit))]

    def combined_history(self, limit=100):
        detailed = self.history(limit=limit)
        if detailed:
            return detailed
        return [{"date": item["imported_at"], "source": item["source_type"], "filename": item["filename"], "rows": item["imported_rows"], "success": item["imported_rows"], "failed": 0, "duration": 0.0, "status": item["result"]} for item in ImportHistoryService.history(limit=limit)]
