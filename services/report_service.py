"""Unified report aggregation, preview, and export service for v1.0."""

import csv
import json
import os
import re
from dataclasses import asdict, dataclass
from datetime import date, datetime, timedelta
from threading import RLock
from time import monotonic

from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from services.action_center import ActionCenterService
from services.analytics_service import AnalyticsService
from services.dashboard_service import DashboardService
from services.opportunity_engine import OpportunityEngine
from services.school_profile_service import SchoolProfileService


REPORT_TYPES = (
    "School Report",
    "Opportunity Report",
    "CRM Report",
    "Weekly Report",
    "Monthly Report",
    "Executive Summary",
)
EXPORT_FORMATS = ("PDF", "Excel", "CSV")


@dataclass(frozen=True)
class ReportSection:
    title: str
    columns: tuple
    rows: tuple

    def to_dict(self):
        return {"title": self.title, "columns": list(self.columns), "rows": [list(row) for row in self.rows]}


@dataclass(frozen=True)
class ReportDocument:
    report_type: str
    title: str
    generated_at: str
    filters: dict
    sections: tuple

    def to_dict(self):
        return {
            "report_type": self.report_type,
            "title": self.title,
            "generated_at": self.generated_at,
            "filters": dict(self.filters),
            "sections": [section.to_dict() for section in self.sections],
        }


class ReportService:
    """Create one query-efficient report model and serialize it consistently."""

    CACHE_TTL_SECONDS = 300
    dashboard_service = DashboardService
    opportunity_engine = OpportunityEngine
    action_service = ActionCenterService
    profile_service = SchoolProfileService
    analytics_service = AnalyticsService
    _cache = {}
    _lock = RLock()

    @classmethod
    def aggregate(cls, report_type, filters=None, force_refresh=False):
        selected_type = str(report_type or "").strip()
        if selected_type not in REPORT_TYPES:
            raise ValueError(f"Unsupported report type: {selected_type}")
        selected_filters = cls._filters(filters)
        cache_key = (selected_type, tuple(sorted(selected_filters.items())))
        now = monotonic()
        with cls._lock:
            cached = cls._cache.get(cache_key)
            if not force_refresh and cached and now - cached["at"] < cls.CACHE_TTL_SECONDS:
                return cached["document"]

        if selected_type == "School Report":
            sections, subject = cls._school_report(selected_filters)
        elif selected_type == "Opportunity Report":
            sections, subject = cls._opportunity_report(selected_filters)
        elif selected_type == "CRM Report":
            sections, subject = cls._crm_report(selected_filters)
        else:
            sections, subject = cls._dashboard_report(selected_type, selected_filters)
        document = ReportDocument(
            report_type=selected_type,
            title=f"EduBid Insight — {selected_type}{subject}",
            generated_at=datetime.now().astimezone().isoformat(timespec="seconds"),
            filters=selected_filters,
            sections=tuple(sections),
        )
        with cls._lock:
            cls._cache[cache_key] = {"at": monotonic(), "document": document}
        return document

    @classmethod
    def preview(cls, report_or_type, filters=None, force_refresh=False):
        document = cls._document(report_or_type, filters, force_refresh)
        lines = [document.title, f"Generated: {document.generated_at}"]
        active_filters = [f"{key}: {value}" for key, value in document.filters.items() if value]
        if active_filters:
            lines.append("Filters: " + " | ".join(active_filters))
        for section in document.sections:
            lines.extend(("", f"[{section.title}]", " | ".join(section.columns)))
            if not section.rows:
                lines.append("No data")
            for row in section.rows:
                lines.append(" | ".join(cls._text(value) for value in row))
        return "\n".join(lines)

    @classmethod
    def export(cls, report_or_type, file_path, export_format=None, filters=None):
        document = cls._document(report_or_type, filters, False)
        selected_format = cls._export_format(export_format, file_path)
        directory = os.path.dirname(os.path.abspath(file_path))
        os.makedirs(directory, exist_ok=True)
        if selected_format == "PDF":
            cls._export_pdf(document, file_path)
        elif selected_format == "Excel":
            cls._export_excel(document, file_path)
        else:
            cls._export_csv(document, file_path)
        return file_path

    @classmethod
    def clear_cache(cls):
        with cls._lock:
            cls._cache.clear()

    @classmethod
    def _school_report(cls, filters):
        school_filter = filters["school"]
        if not school_filter:
            raise ValueError("School Report requires a school filter")
        school = cls.profile_service.resolve_school(school_filter)
        if school is None:
            raise ValueError(f"School was not found: {school_filter}")
        profile = cls.profile_service.get_profile(school["school_code"], school=school)
        if not cls._profile_matches(profile, filters):
            raise ValueError("School does not match the selected region or office")
        context = profile.get("opportunity_context") or {}
        projects = cls._filter_records(
            context.get("projects") or [], filters, "category", ("updated_at", "start_year")
        )
        contracts = cls._filter_records(
            context.get("contracts") or [], filters, "category", ("contract_date", "imported_at")
        )
        crm = cls._filter_records(
            context.get("crm_activities") or [], filters, None, ("activity_date",)
        )
        opportunity = profile.get("opportunity")
        suggested = (
            cls.opportunity_engine.generate_actions(opportunity, persist=False)
            if opportunity else []
        )
        actions = profile.get("actions") or {}
        sections = [
            cls._section("School Profile", ("Field", "Value"), (
                ("School Name", profile["school"].get("school_name", "")),
                ("School Code", profile["school"].get("school_code", "")),
                ("Region", profile["school"].get("region", "")),
                ("Office", school.get("office", "")),
                ("Address", profile["school"].get("address", "")),
                ("Students", profile["school"].get("student_count", 0)),
                ("Classes", profile["school"].get("class_count", 0)),
            )),
            cls._section("Opportunity Score", ("Metric", "Value"), (
                ("Score", opportunity.score if opportunity else 0),
                ("Priority", opportunity.priority if opportunity else "-"),
                ("Confidence", opportunity.confidence if opportunity else "-"),
                ("Recommendation", opportunity.recommendation if opportunity else "-"),
                ("Next Action", opportunity.next_action if opportunity else "-"),
            )),
            cls._section("Evidence", ("Reason",), ((item,) for item in (opportunity.evidence if opportunity else []))),
            cls._section("Projects", ("Project", "Category", "Status", "Budget", "Start", "End"), (
                (item.get("project_name"), item.get("category"), item.get("status"), item.get("budget"), item.get("start_year"), item.get("end_year"))
                for item in projects
            )),
            cls._section("Contracts", ("Date", "Product", "Category", "Vendor", "Amount", "Source"), (
                (item.get("contract_date"), item.get("product"), item.get("category"), item.get("vendor"), item.get("amount"), item.get("source_type"))
                for item in contracts
            )),
            cls._section("CRM History", ("Date", "Type", "Contact", "Status", "Memo"), (
                (item.get("activity_date"), item.get("activity_type"), item.get("contact_person"), item.get("status"), item.get("memo"))
                for item in crm
            )),
            cls._section("Recommended Actions", ("Type", "Title", "Priority", "Due", "Source"), (
                (item.action_type, item.title, item.priority, item.due_date, "Opportunity")
                for item in suggested
            )),
            cls._section("Current Actions", ("Type", "Title", "Status", "Priority", "Due"), (
                (item.action_type, item.title, item.status, item.priority, item.due_date)
                for item in actions.get("current_actions", [])
            )),
            cls._section("Timeline", ("Date", "Source", "Title", "Description"), (
                [
                    (item.get("timestamp"), item.get("source"), item.get("title"), item.get("description"))
                    for item in profile.get("recent_activity", [])
                    if cls._date_matches(item.get("timestamp"), filters)
                ]
                + [
                    (item.updated_at, "CRM Action", item.title, item.status)
                    for item in actions.get("action_timeline", [])
                    if cls._date_matches(item.updated_at, filters)
                ]
            )),
        ]
        analytics = profile.get("analytics") or {}
        sections.extend(cls._chart_sections(analytics.get("project_analytics") or {}))
        return sections, f" — {profile['school'].get('school_name', school_filter)}"

    @classmethod
    def _opportunity_report(cls, filters):
        snapshot = cls.opportunity_engine.dashboard(limit=1000, persist=False, cached_only=True)
        results = cls._filter_opportunities(snapshot.get("all_opportunities", []), filters)
        sections = [
            cls._section("Opportunity Schools", ("School", "Code", "Score", "Priority", "Confidence", "Recommendation"), (
                (item.school_name, item.school_id, item.score, item.priority, item.confidence, item.recommendation)
                for item in results
            )),
            cls._section("Evidence", ("School", "Evidence"), (
                (item.school_name, evidence) for item in results for evidence in item.evidence
            )),
        ]
        return sections, ""

    @classmethod
    def _crm_report(cls, filters):
        actions = cls.action_service.search(
            school=filters["school"], due_from=filters["date_from"], due_to=filters["date_to"]
        )
        activities = cls.action_service.recent_activity(limit=200)
        if filters["school"]:
            selected = filters["school"].casefold()
            actions = [item for item in actions if selected in item.school_id.casefold()]
            activities = [item for item in activities if selected in str(item.get("school_id", "")).casefold()]
        activities = [item for item in activities if cls._date_matches(item.get("timestamp"), filters)]
        sections = [
            cls._section("CRM Actions", ("School", "Type", "Title", "Status", "Priority", "Due", "Completed"), (
                (item.school_id, item.action_type, item.title, item.status, item.priority, item.due_date, item.completed_date)
                for item in actions
            )),
            cls._section("CRM Activity", ("Date", "School", "Type", "Description", "Note"), (
                (item.get("timestamp"), item.get("school_id"), item.get("activity_type"), item.get("description"), item.get("note"))
                for item in activities
            )),
        ]
        return sections, ""

    @classmethod
    def _dashboard_report(cls, report_type, filters):
        snapshot = cls.dashboard_service.get_dashboard()
        priorities = cls._filter_opportunities(snapshot.get("priority_schools", []), filters)
        school_filter = filters["school"].casefold()

        def action_matches(item):
            return (not school_filter or school_filter in item.school_id.casefold()) and cls._date_matches(item.due_date, filters)

        today_actions = [item for item in snapshot.get("today_actions", []) if action_matches(item)]
        alerts = [
            item for item in snapshot.get("alerts", [])
            if not school_filter or school_filter in str(item.get("school_id", "")).casefold()
        ]
        kpis = snapshot.get("weekly_kpi") or {}
        sections = [
            cls._section("KPIs", ("KPI", "Value"), ((key.replace("_", " ").title(), value) for key, value in kpis.items())),
            cls._section("Top Opportunity Schools", ("School", "Score", "Priority", "Recommendation"), (
                (item.school_name, item.score, item.priority, item.recommendation) for item in priorities
            )),
            cls._section("Today's Actions", ("School", "Action", "Due", "Status"), (
                (item.school_id, item.title, item.due_date, item.status) for item in today_actions
            )),
            cls._section("Alerts", ("Severity", "Type", "School", "Message"), (
                (item.get("severity"), item.get("type"), item.get("school_name"), item.get("message")) for item in alerts
            )),
        ]
        sections.extend(cls._chart_sections(snapshot.get("portfolio_analytics") or {}))
        period = cls._period_label(report_type, filters)
        return sections, f" — {period}"

    @classmethod
    def _filter_opportunities(cls, opportunities, filters):
        school_filter = filters["school"].casefold()
        filtered = []
        profiles = {}
        for item in opportunities:
            if school_filter and school_filter not in item.school_id.casefold() and school_filter not in item.school_name.casefold():
                continue
            if (filters["date_from"] or filters["date_to"]) and not cls._date_matches(item.generated_at, filters):
                continue
            if filters["region"] or filters["office"] or filters["category"]:
                profile = profiles.setdefault(item.school_id, cls.profile_service.get_profile(item.school_id))
                if not cls._profile_matches(profile, filters):
                    continue
                if filters["category"]:
                    category = filters["category"].casefold()
                    projects = (profile.get("opportunity_context") or {}).get("projects") or []
                    contracts = (profile.get("opportunity_context") or {}).get("contracts") or []
                    if not any(category == str(row.get("category") or "").casefold() for row in [*projects, *contracts]):
                        continue
            filtered.append(item)
        return sorted(filtered, key=lambda item: (-item.score, item.school_name.casefold()))

    @staticmethod
    def _profile_matches(profile, filters):
        school = profile.get("school") or {}
        if filters["region"] and filters["region"].casefold() not in str(school.get("region") or "").casefold():
            return False
        if filters["office"]:
            office = (profile.get("school_metadata") or {}).get("office", "")
            if filters["office"].casefold() not in str(office).casefold():
                return False
        return True

    @classmethod
    def _filter_records(cls, records, filters, category_field, date_fields):
        category = filters["category"].casefold()
        result = []
        for item in records:
            if category_field and category and category != str(item.get(category_field) or "").casefold():
                continue
            value = next((item.get(field) for field in date_fields if item.get(field)), None)
            if value and not cls._date_matches(value, filters):
                continue
            result.append(item)
        return result

    @staticmethod
    def _chart_sections(analytics):
        return [
            ReportService._section("Chart — Budget Trends", ("Year", "Projects", "Budget"), (
                (row.get("year"), row.get("project_count"), row.get("budget"))
                for row in analytics.get("budget_trends", [])
            )),
            ReportService._section("Chart — Category Distribution", ("Category", "Projects", "Budget"), (
                (row.get("category"), row.get("project_count"), row.get("budget"))
                for row in analytics.get("category_distribution", [])
            )),
        ]

    @staticmethod
    def _section(title, columns, rows):
        return ReportSection(title, tuple(columns), tuple(tuple(row) for row in rows))

    @staticmethod
    def _filters(filters):
        source = filters or {}
        result = {
            "school": str(source.get("school") or "").strip(),
            "region": str(source.get("region") or "").strip(),
            "office": str(source.get("office") or "").strip(),
            "date_from": ReportService._optional_date(source.get("date_from")),
            "date_to": ReportService._optional_date(source.get("date_to")),
            "category": str(source.get("category") or "").strip(),
        }
        if result["date_from"] and result["date_to"] and result["date_from"] > result["date_to"]:
            raise ValueError("date_from must not be after date_to")
        return result

    @staticmethod
    def _date_matches(value, filters):
        selected = ReportService._optional_date(value)
        if selected is None:
            return not filters["date_from"] and not filters["date_to"]
        return (
            (not filters["date_from"] or selected >= filters["date_from"])
            and (not filters["date_to"] or selected <= filters["date_to"])
        )

    @staticmethod
    def _optional_date(value):
        text = str(value or "").strip()[:10]
        if not text:
            return ""
        try:
            return date.fromisoformat(text).isoformat()
        except ValueError as error:
            raise ValueError(f"Invalid date: {value}") from error

    @staticmethod
    def _period_label(report_type, filters):
        if filters["date_from"] or filters["date_to"]:
            return f"{filters['date_from'] or 'Start'} to {filters['date_to'] or 'Today'}"
        today = date.today()
        if report_type == "Weekly Report":
            return f"Week of {(today - timedelta(days=today.weekday())).isoformat()}"
        if report_type == "Monthly Report":
            return today.strftime("%Y-%m")
        return "Executive Overview"

    @staticmethod
    def _document(report_or_type, filters, force_refresh):
        if isinstance(report_or_type, ReportDocument):
            return report_or_type
        return ReportService.aggregate(report_or_type, filters, force_refresh)

    @staticmethod
    def _export_format(export_format, file_path):
        selected = str(export_format or "").strip().casefold()
        extension = os.path.splitext(str(file_path))[1].casefold()
        mapping = {"pdf": "PDF", "excel": "Excel", "xlsx": "Excel", "csv": "CSV"}
        if selected:
            if selected not in mapping:
                raise ValueError(f"Unsupported export format: {export_format}")
            return mapping[selected]
        if extension not in {".pdf", ".xlsx", ".csv"}:
            raise ValueError("Export file must use .pdf, .xlsx, or .csv")
        return {".pdf": "PDF", ".xlsx": "Excel", ".csv": "CSV"}[extension]

    @classmethod
    def _export_csv(cls, document, file_path):
        with open(file_path, "w", encoding="utf-8-sig", newline="") as csv_file:
            writer = csv.writer(csv_file)
            writer.writerow([document.title])
            writer.writerow(["Generated", document.generated_at])
            for section in document.sections:
                writer.writerow([])
                writer.writerow([section.title])
                writer.writerow(section.columns)
                writer.writerows(cls._excel_row(row) for row in section.rows)

    @classmethod
    def _export_excel(cls, document, file_path):
        workbook = Workbook()
        metadata = workbook.active
        metadata.title = "Report"
        metadata.append([document.title])
        metadata.append(["Generated", document.generated_at])
        metadata.append(["Filter", "Value"])
        for key, value in document.filters.items():
            metadata.append([key, value])
        cls._style_sheet(metadata, header_row=3)
        used = {"Report"}
        for section in document.sections:
            title = cls._sheet_title(section.title, used)
            sheet = workbook.create_sheet(title)
            sheet.append(section.columns)
            for row in section.rows:
                sheet.append(cls._excel_row(row))
            cls._style_sheet(sheet)
            if section.title.startswith("Chart") and len(section.rows) > 0 and len(section.columns) >= 3:
                chart = BarChart()
                chart.title = section.title.replace("Chart — ", "")
                chart.add_data(Reference(sheet, min_col=3, min_row=1, max_row=len(section.rows) + 1), titles_from_data=True)
                chart.set_categories(Reference(sheet, min_col=1, min_row=2, max_row=len(section.rows) + 1))
                sheet.add_chart(chart, "E2")
        workbook.save(file_path)
        workbook.close()

    @classmethod
    def _export_pdf(cls, document, file_path):
        lines = cls.preview(document).splitlines()
        pages = [lines[index:index + 31] for index in range(0, len(lines), 31)] or [[]]
        page_count = len(pages)
        font_number = 3 + page_count * 2
        descendant_number = font_number + 1
        kids = " ".join(f"{3 + index * 2} 0 R" for index in range(page_count))
        objects = [
            b"<< /Type /Catalog /Pages 2 0 R >>",
            f"<< /Type /Pages /Kids [{kids}] /Count {page_count} >>".encode("ascii"),
        ]
        for index, page_lines in enumerate(pages):
            content_number = 4 + index * 2
            page_object = (
                f"<< /Type /Page /Parent 2 0 R /MediaBox [0 0 595 842] "
                f"/Resources << /Font << /F1 {font_number} 0 R >> >> "
                f"/Contents {content_number} 0 R >>"
            ).encode("ascii")
            commands = ["BT", "/F1 11 Tf", "45 800 Td"]
            for line_index, line in enumerate(page_lines):
                if line_index:
                    commands.append("0 -23 Td")
                encoded = cls._text(line)[:105].encode("utf-16-be").hex().upper()
                commands.append(f"<{encoded}> Tj")
            commands.append("ET")
            content = "\n".join(commands).encode("ascii")
            objects.extend((
                page_object,
                b"<< /Length " + str(len(content)).encode("ascii") + b" >>\nstream\n" + content + b"\nendstream",
            ))
        objects.extend((
            (
                b"<< /Type /Font /Subtype /Type0 /BaseFont /HYGoThic-Medium "
                b"/Encoding /UniKS-UCS2-H /DescendantFonts ["
                + str(descendant_number).encode("ascii") + b" 0 R] >>"
            ),
            (
                b"<< /Type /Font /Subtype /CIDFontType0 /BaseFont /HYGoThic-Medium "
                b"/CIDSystemInfo << /Registry (Adobe) /Ordering (Korea1) /Supplement 2 >> /DW 1000 >>"
            ),
        ))
        pdf = bytearray(b"%PDF-1.4\n%\xE2\xE3\xCF\xD3\n")
        offsets = [0]
        for number, body in enumerate(objects, start=1):
            offsets.append(len(pdf))
            pdf.extend(f"{number} 0 obj\n".encode("ascii"))
            pdf.extend(body)
            pdf.extend(b"\nendobj\n")
        xref = len(pdf)
        pdf.extend(f"xref\n0 {len(objects) + 1}\n".encode("ascii"))
        pdf.extend(b"0000000000 65535 f \n")
        for offset in offsets[1:]:
            pdf.extend(f"{offset:010d} 00000 n \n".encode("ascii"))
        pdf.extend(
            f"trailer\n<< /Size {len(objects) + 1} /Root 1 0 R >>\nstartxref\n{xref}\n%%EOF\n".encode("ascii")
        )
        with open(file_path, "wb") as pdf_file:
            pdf_file.write(pdf)

    @staticmethod
    def _sheet_title(title, used):
        base = re.sub(r"[\\/*?:\[\]]", "-", title)[:31] or "Section"
        selected = base
        counter = 2
        while selected in used:
            suffix = f" {counter}"
            selected = base[:31 - len(suffix)] + suffix
            counter += 1
        used.add(selected)
        return selected

    @staticmethod
    def _style_sheet(sheet, header_row=1):
        fill = PatternFill(fill_type="solid", fgColor="1F6AA5")
        for cell in sheet[header_row]:
            cell.fill = fill
            cell.font = Font(bold=True, color="FFFFFF")
            cell.alignment = Alignment(horizontal="center")
        sheet.freeze_panes = f"A{header_row + 1}"
        for cells in sheet.columns:
            width = min(max((len(str(cell.value or "")) for cell in cells), default=0) + 3, 48)
            sheet.column_dimensions[get_column_letter(cells[0].column)].width = width

    @staticmethod
    def _excel_row(row):
        return [
            json.dumps(value, ensure_ascii=False) if isinstance(value, (dict, list, tuple)) else value
            for value in row
        ]

    @staticmethod
    def _text(value):
        if value is None:
            return ""
        if isinstance(value, float):
            return f"{value:,.2f}".rstrip("0").rstrip(".")
        if isinstance(value, (dict, list, tuple)):
            return json.dumps(value, ensure_ascii=False)
        return str(value)
