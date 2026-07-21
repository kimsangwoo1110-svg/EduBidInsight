from tkinter import filedialog

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter


def export_school_excel(rows):

    if not rows:
        return

    file_path = filedialog.asksaveasfilename(
        defaultextension=".xlsx",
        filetypes=[("Excel", "*.xlsx")],
        initialfile="학교검색결과.xlsx"
    )

    if not file_path:
        return

    wb = Workbook()
    ws = wb.active
    ws.title = "학교목록"

    headers = [

        "학교코드",
        "학교명",
        "학교급",
        "교육청",
        "지역",
        "주소",
        "홈페이지",
        "AI중점",
        "디지털",
        "공간혁신",
        "그린스마트",
        "학생수",
        "학급수"

    ]

    ws.append(headers)

    fill = PatternFill(
        fill_type="solid",
        fgColor="4F81BD"
    )

    font = Font(
        bold=True,
        color="FFFFFF"
    )

    for cell in ws[1]:
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(
            horizontal="center",
            vertical="center"
        )

    for row in rows:

        homepage = ""

        if row[6]:

            homepage = str(row[6]).strip()

            if homepage.startswith("http://") or homepage.startswith("https://"):
                pass

            elif "." in homepage:
                homepage = "http://" + homepage

            else:
                homepage = ""

        ws.append([

            row[0],      # 학교코드
            row[1],      # 학교명
            row[2],      # 학교급
            row[3],      # 교육청
            row[4],      # 지역
            row[5],      # 주소
            homepage,
            "✅" if row[7] else "",
            "✅" if row[8] else "",
            "✅" if row[9] else "",
            "✅" if row[10] else "",
            row[11],
            row[12]

        ])

    for r in range(2, ws.max_row + 1):

        # 홈페이지 하이퍼링크
        cell = ws.cell(r, 7)

        if cell.value:

            cell.hyperlink = cell.value
            cell.style = "Hyperlink"

        # 학생수
        student = ws.cell(r, 12)

        if isinstance(student.value, int):
            student.number_format = '#,##0'

        # 학급수
        classroom = ws.cell(r, 13)

        if isinstance(classroom.value, int):
            classroom.number_format = '#,##0'

    # 첫 줄 고정
    ws.freeze_panes = "A2"

    # 자동 필터
    ws.auto_filter.ref = ws.dimensions

    # 열 너비 자동
    for column_cells in ws.columns:

        max_length = 0
        column = get_column_letter(column_cells[0].column)

        for cell in column_cells:

            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass

        ws.column_dimensions[column].width = min(max_length + 4, 45)

    wb.save(file_path)