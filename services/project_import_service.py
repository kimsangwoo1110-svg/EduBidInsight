"""Excel bulk registration and update for one school's planned projects."""

from contextlib import closing
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill

from services import database
from services.project_service import ProjectService


PROJECT_IMPORT_FIELDS = (
    "project_name", "category", "status", "budget", "start_year", "end_year",
    "expected_procurement_date", "memo", "source",
)
PROJECT_IMPORT_LABELS = {
    "project_name": "사업명",
    "category": "사업 분류",
    "status": "상태",
    "budget": "예산",
    "start_year": "시작 연도",
    "end_year": "종료 연도",
    "expected_procurement_date": "예상 조달일",
    "memo": "메모",
    "source": "출처",
}
PROJECT_IMPORT_ALIASES = {
    "project_name": ("사업명", "프로젝트명", "예정사업명", "사업 이름", "project name"),
    "category": ("사업 분류", "분류", "카테고리", "category"),
    "status": ("상태", "사업 상태", "진행 상태", "status"),
    "budget": ("예산", "사업비", "예상 예산", "budget"),
    "start_year": ("시작 연도", "시작년도", "시작년", "start year"),
    "end_year": ("종료 연도", "종료년도", "종료년", "end year"),
    "expected_procurement_date": (
        "예상 조달일", "조달 예정일", "예정 조달일", "구매 예정일", "procurement date",
    ),
    "memo": ("메모", "비고", "내용", "memo", "note"),
    "source": ("출처", "데이터 출처", "source"),
}


class ProjectImportService:
    @classmethod
    def import_excel(cls, file_path, school_code):
        path = Path(file_path).resolve()
        if path.suffix.lower() != ".xlsx" or not path.is_file():
            raise ValueError(".xlsx 형식의 예정사업 파일을 선택하세요.")
        if not str(school_code or "").strip():
            raise ValueError("학교코드가 필요합니다.")

        workbook = load_workbook(path, read_only=True, data_only=True)
        try:
            sheet = workbook.active
            rows = sheet.iter_rows(values_only=True)
            headers = [str(value or "").strip() for value in next(rows, ())]
            mapping = cls.auto_map(headers)
            if not mapping["project_name"]:
                raise ValueError("필수 열 '사업명'을 찾을 수 없습니다.")
            source_rows = list(rows)
        finally:
            workbook.close()

        existing = {
            cls._identity(project["project_name"]): project
            for project in ProjectService.list_for_school(school_code)
        }
        summary = {"rows": len(source_rows), "inserted": 0, "updated": 0, "failed": 0, "errors": []}
        with closing(database.get_connection()) as connection:
            try:
                connection.execute("BEGIN")
                for row_number, row in enumerate(source_rows, start=2):
                    source = {
                        header: row[index] if index < len(row) else None
                        for index, header in enumerate(headers)
                        if header
                    }
                    if not any(value not in (None, "") for value in source.values()):
                        continue
                    try:
                        values = cls._normalize(source, mapping)
                        identity = cls._identity(values["project_name"])
                        current = existing.get(identity)
                        if current:
                            ProjectService.update(
                                current["id"], **values, connection=connection, commit=False
                            )
                            current.update(values)
                            summary["updated"] += 1
                        else:
                            project_id = ProjectService.create(
                                school_code, **values, connection=connection, commit=False
                            )
                            existing[identity] = {"id": project_id, **values}
                            summary["inserted"] += 1
                    except (TypeError, ValueError) as error:
                        summary["failed"] += 1
                        summary["errors"].append(f"{row_number}행: {error}")
                connection.commit()
            except BaseException:
                connection.rollback()
                raise
        return summary

    @classmethod
    def auto_map(cls, headers):
        normalized = {cls._identity(header): header for header in headers}
        return {
            field: next(
                (normalized[cls._identity(alias)] for alias in aliases if cls._identity(alias) in normalized),
                "",
            )
            for field, aliases in PROJECT_IMPORT_ALIASES.items()
        }

    @classmethod
    def _normalize(cls, source, mapping):
        value = lambda field: source.get(mapping.get(field), "") if mapping.get(field) else ""
        project_name = str(value("project_name") or "").strip()
        if not project_name:
            raise ValueError("사업명이 비어 있습니다.")
        status = str(value("status") or "예정").strip()
        if status not in ProjectService.STATUS_OPTIONS:
            raise ValueError(f"지원하지 않는 상태입니다: {status}")
        budget_text = str(value("budget") or "0").replace(",", "").replace("원", "").strip()
        try:
            budget = float(budget_text or 0)
        except ValueError as error:
            raise ValueError("예산은 숫자여야 합니다.") from error
        if budget < 0:
            raise ValueError("예산은 0 이상이어야 합니다.")
        start_year = cls._year(value("start_year"), "시작 연도")
        end_year = cls._year(value("end_year"), "종료 연도")
        if start_year and end_year and end_year < start_year:
            raise ValueError("종료 연도는 시작 연도보다 빠를 수 없습니다.")
        return {
            "project_name": project_name,
            "category": str(value("category") or "").strip(),
            "status": status,
            "budget": budget,
            "start_year": start_year,
            "end_year": end_year,
            "expected_procurement_date": ProjectService.normalize_date(value("expected_procurement_date")),
            "memo": str(value("memo") or "").strip(),
            "source": str(value("source") or "Excel 일괄 가져오기").strip(),
        }

    @staticmethod
    def create_template(destination):
        workbook = Workbook(); sheet = workbook.active; sheet.title = "예정사업"
        sheet.append([PROJECT_IMPORT_LABELS[field] for field in PROJECT_IMPORT_FIELDS])
        sheet.append(["미래교실 구축", "디지털 교육", "예정", 100000000, 2026, 2027, "2026-11-01", "담당자 협의 예정", "교육청"])
        sheet.freeze_panes = "A2"; sheet.auto_filter.ref = sheet.dimensions
        for cell in sheet[1]:
            cell.font = Font(name="맑은 고딕", bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="4F78A6")
            cell.alignment = Alignment(horizontal="center")
        widths = (24, 16, 12, 16, 12, 12, 16, 32, 16)
        for index, width in enumerate(widths, start=1):
            sheet.column_dimensions[chr(64 + index)].width = width
        workbook.save(destination); workbook.close()
        return destination

    @staticmethod
    def _identity(value):
        return "".join(character for character in str(value or "").casefold() if character.isalnum())

    @staticmethod
    def _year(value, label):
        if value in (None, ""):
            return None
        try:
            year = int(value)
        except (TypeError, ValueError) as error:
            raise ValueError(f"{label}는 숫자여야 합니다.") from error
        if not 2000 <= year <= 2200:
            raise ValueError(f"{label} 범위는 2000~2200입니다.")
        return year
