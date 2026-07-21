"""Excel and CSV connector for education contract files."""

import csv
import io
import os

from openpyxl import load_workbook

from services.connectors.base import BaseConnector
from services.contract_service import ContractService, REQUIRED_FIELDS


CONTRACT_COLUMNS = (
    "school_code",
    "school_name",
    "contract_date",
    "product",
    "category",
    "vendor",
    "quantity",
    "amount",
)
FIELD_LABELS = {
    "school_code": "학교코드",
    "school_name": "학교명",
    "contract_date": "계약일",
    "product": "품목",
    "category": "카테고리",
    "vendor": "업체",
    "quantity": "수량",
    "amount": "금액",
}
COLUMN_ALIASES = {
    "school_code": ("학교코드", "표준학교코드", "schoolcode", "schoolid"),
    "school_name": ("학교명", "학교이름", "schoolname"),
    "contract_date": ("계약일", "계약일자", "계약날짜", "contractdate", "date"),
    "product": ("품목", "품목명", "제품", "제품명", "물품명", "product", "item"),
    "category": ("분류", "카테고리", "category", "type"),
    "vendor": ("업체", "업체명", "공급업체", "계약업체", "vendor", "supplier"),
    "quantity": ("수량", "quantity", "qty"),
    "amount": ("금액", "계약금액", "총액", "amount", "price", "total"),
}


class ContractImportConnector(BaseConnector):
    """Import a mapped worksheet or CSV into the contracts table."""

    source = "계약 파일 가져오기"

    def __init__(self, file_path, sheet_name=None, mapping=None):
        self.file_path = os.path.abspath(file_path)
        self.extension = os.path.splitext(self.file_path)[1].lower()
        if self.extension not in {".xlsx", ".csv"}:
            raise ValueError("only .xlsx and .csv contract files are supported")
        if not os.path.isfile(self.file_path):
            raise ValueError("contract file does not exist")
        available_sheets = self.sheet_names(self.file_path)
        self.sheet_name = sheet_name or available_sheets[0]
        if self.sheet_name not in available_sheets:
            raise ValueError(f"unknown sheet: {self.sheet_name}")
        headers = self.headers(self.file_path, self.sheet_name)
        self.mapping = mapping or self.auto_map(headers)
        self.validate_mapping(self.mapping, headers)
        super().__init__()

    def fetch(self):
        if self.extension == ".csv":
            yield from self._csv_rows(self.file_path)
            return
        workbook = load_workbook(self.file_path, read_only=True, data_only=True)
        try:
            worksheet = workbook[self.sheet_name]
            rows = worksheet.iter_rows(values_only=True)
            headers = [str(value or "").strip() for value in next(rows, ())]
            for values in rows:
                yield {
                    header: values[index] if index < len(values) else None
                    for index, header in enumerate(headers)
                    if header
                }
        finally:
            workbook.close()

    def transform(self, source_row):
        values = {
            field: source_row.get(header, "") if header else ""
            for field, header in self.mapping.items()
        }
        values["source_file"] = os.path.basename(self.file_path)
        return ContractService.validate(values)

    def load(self, contract):
        contract_id = ContractService.save(**contract)
        return "inserted" if contract_id is not None else "skipped"

    def preview(self, limit=20):
        """Return mapped preview rows with validation and duplicate messages."""
        preview_rows = []
        records = self.fetch()
        try:
            for row_number, source_row in enumerate(records, start=2):
                if len(preview_rows) >= limit:
                    break
                try:
                    contract = self.transform(source_row)
                    error = (
                        "중복 계약"
                        if ContractService.duplicate_check(**contract)
                        else ""
                    )
                except (TypeError, ValueError) as validation_error:
                    contract = {
                        field: source_row.get(self.mapping.get(field, ""), "")
                        for field in CONTRACT_COLUMNS
                    }
                    error = str(validation_error)
                preview_rows.append(
                    {"row_number": row_number, "contract": contract, "error": error}
                )
        finally:
            close_records = getattr(records, "close", None)
            if close_records:
                close_records()
        return preview_rows

    @classmethod
    def sheet_names(cls, file_path):
        extension = os.path.splitext(file_path)[1].lower()
        if extension == ".csv":
            return ["CSV"]
        if extension != ".xlsx":
            raise ValueError("only .xlsx and .csv contract files are supported")
        workbook = load_workbook(file_path, read_only=True, data_only=True)
        try:
            return list(workbook.sheetnames)
        finally:
            workbook.close()

    @classmethod
    def headers(cls, file_path, sheet_name=None):
        extension = os.path.splitext(file_path)[1].lower()
        if extension == ".csv":
            reader = csv.DictReader(io.StringIO(cls._read_csv_content(file_path)))
            return list(reader.fieldnames or [])
        workbook = load_workbook(file_path, read_only=True, data_only=True)
        try:
            worksheet = workbook[sheet_name or workbook.sheetnames[0]]
            values = next(worksheet.iter_rows(values_only=True), ())
            return [str(value or "").strip() for value in values if str(value or "").strip()]
        finally:
            workbook.close()

    @classmethod
    def auto_map(cls, headers):
        """Map canonical contract fields using normalized header aliases."""
        normalized_headers = {cls._normalize_header(header): header for header in headers}
        mapping = {}
        used_headers = set()
        for field in CONTRACT_COLUMNS:
            selected = ""
            for alias in COLUMN_ALIASES[field]:
                candidate = normalized_headers.get(cls._normalize_header(alias), "")
                if candidate and candidate not in used_headers:
                    selected = candidate
                    used_headers.add(candidate)
                    break
            mapping[field] = selected
        return mapping

    @staticmethod
    def validate_mapping(mapping, headers):
        available_headers = set(headers)
        missing = [field for field in REQUIRED_FIELDS if not mapping.get(field)]
        if missing:
            raise ValueError(f"required column mappings are missing: {', '.join(missing)}")
        invalid = [header for header in mapping.values() if header and header not in available_headers]
        if invalid:
            raise ValueError(f"mapped columns do not exist: {', '.join(invalid)}")
        selected = [header for header in mapping.values() if header]
        if len(selected) != len(set(selected)):
            raise ValueError("one source column cannot be mapped more than once")
        return True

    @staticmethod
    def _normalize_header(value):
        return "".join(
            character
            for character in str(value or "").strip().casefold()
            if character.isalnum()
        )

    @staticmethod
    def _csv_rows(file_path):
        yield from csv.DictReader(
            io.StringIO(ContractImportConnector._read_csv_content(file_path))
        )

    @staticmethod
    def _read_csv_content(file_path):
        content = None
        last_error = None
        for encoding in ("utf-8-sig", "cp949"):
            try:
                with open(file_path, encoding=encoding, newline="") as csv_file:
                    content = csv_file.read()
                break
            except UnicodeDecodeError as error:
                last_error = error
        if content is None:
            raise ValueError("CSV encoding must be UTF-8 or CP949") from last_error
        return content
