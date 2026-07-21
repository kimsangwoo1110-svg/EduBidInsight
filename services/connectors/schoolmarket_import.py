"""SchoolMarket Excel/CSV import for the Smart Import Framework."""

import csv
import hashlib
import io
import os
import threading
from collections import Counter
from contextlib import closing
from datetime import datetime
from time import perf_counter

from openpyxl import load_workbook

from services import database
from services.base_import import BaseImport, ImportCancelled
from services.contract_service import ContractService
from services.school_service import SchoolService


SCHOOLMARKET_COLUMNS = (
    "contract_number",
    "school_name",
    "purchase_date",
    "product_name",
    "vendor",
    "quantity",
    "amount",
)
SCHOOLMARKET_REQUIRED_FIELDS = (
    "school_name",
    "purchase_date",
    "product_name",
    "vendor",
    "amount",
)
SCHOOLMARKET_FIELD_LABELS = {
    "contract_number": "계약번호",
    "school_name": "학교명",
    "purchase_date": "구매일",
    "product_name": "구매제품",
    "vendor": "판매업체",
    "quantity": "수량",
    "amount": "구매금액",
}
SCHOOLMARKET_COLUMN_ALIASES = {
    "contract_number": (
        "계약번호", "계약 번호", "주문번호", "거래번호", "contractnumber", "contractno", "orderno",
    ),
    "school_name": (
        "학교명", "학교", "기관명", "수요기관명", "구매기관", "schoolname", "institutionname",
    ),
    "purchase_date": (
        "구매일", "구매일자", "계약일", "계약일자", "거래일", "purchase date", "purchasedate", "contractdate",
    ),
    "product_name": (
        "제품명", "상품명", "품목명", "물품명", "구매제품", "productname", "product", "itemname",
    ),
    "vendor": (
        "업체", "업체명", "판매업체", "공급업체", "계약업체", "vendor", "supplier", "seller",
    ),
    "quantity": ("수량", "구매수량", "quantity", "qty"),
    "amount": (
        "금액", "구매금액", "계약금액", "총액", "결제금액", "amount", "totalamount", "price",
    ),
}

CATEGORY_TERMS = (
    ("ICT", ("interactive display", "interactive panel", "전자칠판", "스마트칠판")),
    ("ICT", ("electronic whiteboard",)),
    ("ICT", ("notebook", "laptop", "노트북")),
    ("ICT", ("tablet", "태블릿")),
    ("ICT", ("desktop", "데스크톱", "데스크탑")),
    ("AV", ("projector", "프로젝터", "빔프로젝터")),
    ("AV", ("display", "디스플레이", "모니터")),
    ("AV", ("audio", "오디오", "스피커", "마이크")),
    ("Furniture", ("desk", "책상")),
    ("Furniture", ("chair", "의자")),
    ("Furniture", ("cabinet", "캐비닛", "수납장")),
)


def classify_product(product_name):
    """Classify a SchoolMarket product using deterministic first-match rules."""
    normalized = " ".join(str(product_name or "").casefold().split())
    for category, terms in CATEGORY_TERMS:
        if any(term in normalized for term in terms):
            return category
    return "Other"


class SchoolMarketImport(BaseImport):
    """Import SchoolMarket purchases as contracts with source-aware deduplication."""

    source = "SchoolMarket"

    def __init__(
        self,
        filename,
        sheet_name=None,
        mapping=None,
        history_service=None,
        cancel_event=None,
    ):
        super().__init__(self.source, filename, history_service)
        self.extension = os.path.splitext(self.filename)[1].lower()
        if self.extension not in {".xlsx", ".csv"}:
            raise ValueError("only .xlsx and .csv SchoolMarket files are supported")
        if not os.path.isfile(self.filename):
            raise ValueError("SchoolMarket file does not exist")
        sheets = self.sheet_names(self.filename)
        self.sheet_name = sheet_name or sheets[0]
        headers = self.headers(self.filename, self.sheet_name)
        self.mapping = mapping or self.auto_map(headers)
        self.validate_mapping(self.mapping, headers)
        self.cancel_event = cancel_event or threading.Event()
        self._loaded = None

    def load(self):
        if self._loaded is None:
            self._loaded = list(self._rows())
        return self._loaded

    def validate(self):
        validated = []
        for row_number, source_row in enumerate(self.load(), start=2):
            try:
                purchase = self._transform(source_row)
                validated.append({"row_number": row_number, "purchase": purchase, "error": ""})
            except (TypeError, ValueError) as error:
                validated.append({"row_number": row_number, "purchase": None, "error": str(error)})
        return validated

    def preview(self, limit=100):
        preview_rows = []
        for row in self.validate()[: min(100, max(0, int(limit)))]:
            purchase = row["purchase"] or self._raw_contract(self.load()[row["row_number"] - 2])
            match = SchoolService.match_name(purchase.get("school_name"))
            warning = "" if match else f"학교 미일치: {purchase.get('school_name', '')}"
            preview_rows.append(
                {
                    "row_number": row["row_number"],
                    "contract": {
                        **purchase,
                        "school_code": match["school_code"] if match else "",
                    },
                    "error": row["error"] or warning,
                    "missing_fields": self._missing_fields(purchase),
                    "category": purchase.get("category", "Other"),
                    "school_matched": bool(match),
                }
            )
        return preview_rows

    def save(self, progress_callback=None):
        rows = self.validate()
        total = len(rows)
        result = {
            "imported": 0,
            "skipped": 0,
            "duplicates": 0,
            "warnings": [],
            "school_matches": 0,
            "school_rows": 0,
            "category_summary": Counter(),
        }
        with closing(database.get_connection()) as connection:
            try:
                connection.execute("BEGIN")
                for processed, row in enumerate(rows, start=1):
                    if self.cancel_event.is_set():
                        raise ImportCancelled()
                    if row["error"]:
                        result["skipped"] += 1
                        result["warnings"].append(f"Row {row['row_number']}: {row['error']}")
                    else:
                        self._save_purchase(row, result, connection)
                    if progress_callback:
                        progress_callback(
                            stage="Importing...",
                            processed=processed,
                            total=total,
                            percentage=round(processed * 100 / total) if total else 100,
                        )
                if self.cancel_event.is_set():
                    raise ImportCancelled()
                if progress_callback:
                    progress_callback(stage="Saving...", processed=total, total=total, percentage=100)
                connection.commit()
            except BaseException:
                connection.rollback()
                raise
        result["category_summary"] = dict(sorted(result["category_summary"].items()))
        return result

    def run(self, progress_callback=None):
        started = datetime.now().astimezone()
        timer = perf_counter()
        result = self._empty_result()
        status = "FAILED"
        exception_text = ""
        try:
            self._notify(progress_callback, "Reading file...", 0, 0, 0)
            self.load()
            self._notify(progress_callback, "Validating...", 0, len(self._loaded), 0)
            result = self.save(progress_callback)
            status = "PARTIAL" if result["skipped"] or result["warnings"] else "SUCCESS"
        except ImportCancelled:
            status = "Cancelled"
            result = self._empty_result()
            result["warnings"] = ["Import cancelled by user; all changes were rolled back."]
        except Exception as error:
            exception_text = f"{type(error).__name__}: {error}"
            result["warnings"] = [exception_text]
            self.logger.exception("%s import failed", self.source)
        elapsed = perf_counter() - timer
        finished = datetime.now().astimezone()
        match_rate = (
            result["school_matches"] / result["school_rows"] * 100
            if result["school_rows"] else 0.0
        )
        self.log(status, result["imported"])
        summary = {
            "status": status,
            "started_at": started.isoformat(timespec="seconds"),
            "finished_at": finished.isoformat(timespec="seconds"),
            "elapsed": elapsed,
            **result,
            "failed": result["skipped"],
            "school_match_rate": match_rate,
            "cancelled": status == "Cancelled",
            "exception": exception_text,
        }
        self.logger.info(
            "%s import finished | imported=%d | skipped=%d | duplicates=%d | "
            "warnings=%d | school_matches=%d/%d | categories=%s | elapsed=%.3f",
            self.source, result["imported"], result["skipped"], result["duplicates"],
            len(result["warnings"]), result["school_matches"], result["school_rows"],
            result["category_summary"], elapsed,
        )
        return summary

    def cancel(self):
        self.cancel_event.set()

    def _save_purchase(self, row, result, connection):
        purchase = row["purchase"]
        result["school_rows"] += 1
        match = SchoolService.match_name(purchase["school_name"], connection=connection)
        if match:
            result["school_matches"] += 1
            school_code = match["school_code"]
            school_name = match["school_name"]
        else:
            school_name = purchase["school_name"]
            school_code = self._unmatched_school_code(school_name)
            result["warnings"].append(f"Row {row['row_number']}: school not matched: {school_name}")
        contract = ContractService.validate(
            {
                "school_code": school_code,
                "school_name": school_name,
                "contract_date": purchase["contract_date"],
                "product": purchase["product"],
                "category": purchase["category"],
                "vendor": purchase["vendor"],
                "quantity": purchase["quantity"],
                "amount": purchase["amount"],
                "source_file": os.path.basename(self.filename),
            }
        )
        record_key = self._record_key(purchase, contract)
        duplicate = database.schoolmarket_key_exists(record_key, connection=connection)
        if not purchase["contract_number"] and not duplicate:
            duplicate = database.contract_duplicate_exists(contract, connection=connection)
        if duplicate:
            result["duplicates"] += 1
            return
        contract_id = database.add_contract(contract, connection=connection, commit=False)
        database.add_schoolmarket_key(
            record_key, purchase["contract_number"], contract_id, connection
        )
        result["imported"] += 1
        result["category_summary"][purchase["category"]] += 1

    def _transform(self, source_row):
        values = {
            field: source_row.get(header, "") if header else ""
            for field, header in self.mapping.items()
        }
        purchase = {
            "contract_number": str(values.get("contract_number") or "").strip(),
            "school_name": str(values.get("school_name") or "").strip(),
            "contract_date": ContractService.normalize_date(values.get("purchase_date")),
            "product": str(values.get("product_name") or "").strip(),
            "vendor": str(values.get("vendor") or "").strip(),
            "quantity": ContractService.normalize_quantity(values.get("quantity")),
            "amount": ContractService.normalize_amount(values.get("amount")),
        }
        purchase["category"] = classify_product(purchase["product"])
        missing = [
            field for field in ("school_name", "contract_date", "product", "vendor", "amount")
            if purchase.get(field) in (None, "")
        ]
        if missing:
            raise ValueError(f"required fields are missing: {', '.join(missing)}")
        return purchase

    def _raw_contract(self, source_row):
        values = {
            field: source_row.get(header, "") if header else ""
            for field, header in self.mapping.items()
        }
        product = str(values.get("product_name") or "").strip()
        return {
            "school_name": str(values.get("school_name") or "").strip(),
            "contract_date": values.get("purchase_date", ""),
            "product": product,
            "vendor": values.get("vendor", ""),
            "quantity": values.get("quantity", ""),
            "amount": values.get("amount", ""),
            "category": classify_product(product),
        }

    @staticmethod
    def _missing_fields(purchase):
        return [
            field for field in ("school_name", "contract_date", "product", "vendor", "amount")
            if purchase.get(field) in (None, "")
        ]

    @staticmethod
    def _unmatched_school_code(school_name):
        digest = hashlib.sha1(str(school_name).encode("utf-8")).hexdigest()[:12]
        return f"UNMATCHED-{digest}"

    @staticmethod
    def _record_key(purchase, contract):
        number = "".join(purchase["contract_number"].split()).casefold()
        if number:
            return f"number:{number}"
        composite = "|".join(
            str(contract[field] if field in contract else "").strip().casefold()
            for field in ("school_code", "contract_date", "product", "vendor", "quantity", "amount")
        )
        return "composite:" + hashlib.sha256(composite.encode("utf-8")).hexdigest()

    def _rows(self):
        if self.extension == ".csv":
            yield from csv.DictReader(io.StringIO(self._read_csv_content(self.filename)))
            return
        workbook = load_workbook(self.filename, read_only=True, data_only=True)
        try:
            worksheet = workbook[self.sheet_name]
            rows = worksheet.iter_rows(values_only=True)
            headers = [str(value or "").strip() for value in next(rows, ())]
            for values in rows:
                yield {
                    header: values[index] if index < len(values) else None
                    for index, header in enumerate(headers) if header
                }
        finally:
            workbook.close()

    @classmethod
    def sheet_names(cls, file_path):
        extension = os.path.splitext(file_path)[1].lower()
        if extension == ".csv":
            return ["CSV"]
        if extension != ".xlsx":
            raise ValueError("only .xlsx and .csv SchoolMarket files are supported")
        workbook = load_workbook(file_path, read_only=True, data_only=True)
        try:
            return list(workbook.sheetnames)
        finally:
            workbook.close()

    @classmethod
    def headers(cls, file_path, sheet_name=None):
        extension = os.path.splitext(file_path)[1].lower()
        if extension == ".csv":
            return list(csv.DictReader(io.StringIO(cls._read_csv_content(file_path))).fieldnames or [])
        workbook = load_workbook(file_path, read_only=True, data_only=True)
        try:
            worksheet = workbook[sheet_name or workbook.sheetnames[0]]
            values = next(worksheet.iter_rows(values_only=True), ())
            return [str(value or "").strip() for value in values if str(value or "").strip()]
        finally:
            workbook.close()

    @classmethod
    def auto_map(cls, headers):
        normalized = {cls._normalize(header): header for header in headers}
        return {
            field: next(
                (normalized[cls._normalize(alias)] for alias in SCHOOLMARKET_COLUMN_ALIASES[field]
                 if cls._normalize(alias) in normalized),
                "",
            )
            for field in SCHOOLMARKET_COLUMNS
        }

    @staticmethod
    def validate_mapping(mapping, headers):
        missing = [field for field in SCHOOLMARKET_REQUIRED_FIELDS if not mapping.get(field)]
        if missing:
            raise ValueError(f"required column mappings are missing: {', '.join(missing)}")
        invalid = [value for value in mapping.values() if value and value not in set(headers)]
        if invalid:
            raise ValueError(f"mapped columns do not exist: {', '.join(invalid)}")
        selected = [value for value in mapping.values() if value]
        if len(selected) != len(set(selected)):
            raise ValueError("one source column cannot be mapped more than once")
        return True

    @staticmethod
    def _normalize(value):
        return "".join(character for character in str(value or "").casefold() if character.isalnum())

    @staticmethod
    def _read_csv_content(file_path):
        for encoding in ("utf-8-sig", "cp949"):
            try:
                with open(file_path, encoding=encoding, newline="") as csv_file:
                    return csv_file.read()
            except UnicodeDecodeError:
                continue
        raise ValueError("CSV encoding must be UTF-8 or CP949")

    @staticmethod
    def _empty_result():
        return {
            "imported": 0, "skipped": 0, "duplicates": 0, "warnings": [],
            "school_matches": 0, "school_rows": 0, "category_summary": {},
        }

    @staticmethod
    def _notify(callback, stage, processed, total, percentage):
        if callback:
            callback(stage=stage, processed=processed, total=total, percentage=percentage)


SchoolMarketImportConnector = SchoolMarketImport


class SchoolMarketService:
    """Read-only SchoolMarket purchase identity service."""

    @staticmethod
    def contract_ids_for_school(school_code):
        return database.find_schoolmarket_contract_ids(school_code)
