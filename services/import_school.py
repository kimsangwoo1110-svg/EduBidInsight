from contextlib import closing

from openpyxl import load_workbook

from services.database import add_school, get_connection, school_exists
from services.region_data import normalize_office, normalize_region


def value(cell_value):
    return str(cell_value or "").strip()


def integer_value(cell_value):
    if cell_value in (None, ""):
        return 0
    if isinstance(cell_value, bool):
        return int(cell_value)
    return int(str(cell_value).replace(",", "").strip())


def flag_value(cell_value):
    return int(str(cell_value or "").strip().lower() in {"1", "true", "y", "yes", "o", "✓"})


def import_school_data(file_path):
    """Import schools from the export-compatible Excel layout.

    The returned summary contains inserted, updated, skipped, and error counts.
    Invalid rows never stop the remaining rows from being processed.
    """
    summary = {"inserted": 0, "updated": 0, "skipped": 0, "errors": 0}
    seen_school_codes = set()
    workbook = load_workbook(file_path, read_only=True, data_only=True)
    worksheet = workbook.active

    try:
        with closing(get_connection()) as connection, connection:
            for row in worksheet.iter_rows(min_row=2, values_only=True):
                school_code = value(row[0] if len(row) > 0 else "")
                school_name = value(row[1] if len(row) > 1 else "")

                if not school_code or not school_name:
                    summary["skipped"] += 1
                    continue

                if school_code in seen_school_codes:
                    summary["skipped"] += 1
                    continue
                seen_school_codes.add(school_code)

                try:
                    is_update = school_exists(school_code, connection)
                    add_school(
                        school_code=school_code,
                        name=school_name,
                        school_type=value(row[2] if len(row) > 2 else ""),
                        office=normalize_office(value(row[3] if len(row) > 3 else "")),
                        region=normalize_region(value(row[4] if len(row) > 4 else "")),
                        address=value(row[5] if len(row) > 5 else ""),
                        homepage=value(row[6] if len(row) > 6 else ""),
                        ai_school=flag_value(row[7] if len(row) > 7 else ""),
                        digital_school=flag_value(row[8] if len(row) > 8 else ""),
                        space_innovation=flag_value(row[9] if len(row) > 9 else ""),
                        green_smart=flag_value(row[10] if len(row) > 10 else ""),
                        student_count=integer_value(row[11] if len(row) > 11 else ""),
                        class_count=integer_value(row[12] if len(row) > 12 else ""),
                        connection=connection,
                        commit=False,
                    )
                    summary["updated" if is_update else "inserted"] += 1
                except Exception:
                    summary["errors"] += 1
    finally:
        workbook.close()

    return summary
