import os
import tempfile
import unittest

from openpyxl import Workbook, load_workbook

from services import database
from services.project_import_service import ProjectImportService
from services.project_service import ProjectService


class ProjectManagementTest(unittest.TestCase):
    def setUp(self):
        self.temporary = tempfile.TemporaryDirectory()
        self.original_database = database.DB_NAME
        database.DB_NAME = os.path.join(self.temporary.name, "projects.db")
        database.create_database()

    def tearDown(self):
        database.DB_NAME = self.original_database
        self.temporary.cleanup()

    def workbook(self, name, headers, rows):
        path = os.path.join(self.temporary.name, name)
        workbook = Workbook()
        sheet = workbook.active
        sheet.append(headers)
        for row in rows:
            sheet.append(row)
        workbook.save(path)
        workbook.close()
        return path

    def test_project_crud_preserves_management_fields(self):
        project_id = ProjectService.create(
            "S-26", "미래교실 구축", "교육환경", "예정", 120_000_000,
            2026, 2027, "사용자", "2026.11.01", "사전 협의 필요",
        )

        project = ProjectService.list_for_school("S-26")[0]
        self.assertEqual(project["id"], project_id)
        self.assertEqual(project["expected_procurement_date"], "2026-11-01")
        self.assertEqual(project["memo"], "사전 협의 필요")

        ProjectService.update(
            project_id, "미래교실 구축", "교육환경", "진행중", 130_000_000,
            2026, 2027, "사용자", "2026/12/01", "예산 확정",
        )
        updated = ProjectService.list_for_school("S-26")[0]
        self.assertEqual(updated["status"], "진행중")
        self.assertEqual(updated["budget"], 130_000_000)
        self.assertEqual(updated["expected_procurement_date"], "2026-12-01")
        self.assertEqual(updated["memo"], "예산 확정")

        self.assertTrue(ProjectService.delete(project_id))
        self.assertEqual(ProjectService.list_for_school("S-26"), [])

    def test_excel_import_inserts_updates_and_reports_invalid_rows(self):
        source = self.workbook(
            "예정사업.xlsx",
            ["예정사업명", "분류", "사업 상태", "사업비", "시작연도", "종료연도",
             "조달 예정일", "비고", "데이터 출처"],
            [
                ["AI 교실", "AI 교육", "예정", "100,000,000원", 2026, 2027,
                 "2026-10-15", "담당자 협의", "교육청"],
                ["오류 사업", "시설", "알 수 없음", 1000, 2026, 2026,
                 "2026-12-01", "", "사용자"],
            ],
        )

        result = ProjectImportService.import_excel(source, "S-26")
        self.assertEqual(result["rows"], 2)
        self.assertEqual(result["inserted"], 1)
        self.assertEqual(result["failed"], 1)
        self.assertIn("3행", result["errors"][0])

        update = self.workbook(
            "예정사업_수정.xlsx",
            ["사업명", "상태", "예산", "예상 조달일", "메모"],
            [["AI 교실", "진행중", 150_000_000, "2026.11.20", "계약 준비"]],
        )
        result = ProjectImportService.import_excel(update, "S-26")
        self.assertEqual(result["updated"], 1)
        project = ProjectService.list_for_school("S-26")[0]
        self.assertEqual(project["budget"], 150_000_000)
        self.assertEqual(project["expected_procurement_date"], "2026-11-20")
        self.assertEqual(project["memo"], "계약 준비")

    def test_template_contains_official_korean_headers(self):
        destination = os.path.join(self.temporary.name, "양식.xlsx")
        ProjectImportService.create_template(destination)
        workbook = load_workbook(destination, read_only=True)
        try:
            headers = [cell.value for cell in workbook.active[1]]
        finally:
            workbook.close()
        self.assertEqual(headers[0], "사업명")
        self.assertIn("예상 조달일", headers)
        self.assertIn("메모", headers)


if __name__ == "__main__":
    unittest.main()
