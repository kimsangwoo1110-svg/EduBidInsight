import os
import tempfile
import unittest

from openpyxl import Workbook, load_workbook

from services import database
from services.import_center import (
    CRMImport, ImportRunStore, PROFILES, SchoolImport,
    export_failed_rows, failed_rows,
)


class ImportCenterTest(unittest.TestCase):
    def setUp(self):
        self.temporary = tempfile.TemporaryDirectory()
        self.original_database = database.DB_NAME
        database.DB_NAME = os.path.join(self.temporary.name, "import-center.db")
        database.create_database()

    def tearDown(self):
        database.DB_NAME = self.original_database
        self.temporary.cleanup()

    def workbook(self, filename, headers, rows):
        path = os.path.join(self.temporary.name, filename)
        workbook = Workbook()
        sheet = workbook.active
        sheet.append(headers)
        for row in rows:
            sheet.append(row)
        workbook.save(path)
        workbook.close()
        return path

    def test_school_alias_mapping_validation_import_and_failed_export(self):
        source = self.workbook(
            "schools.xlsx",
            ["기관코드", "학교 이름", "학교급", "학생 수"],
            [["S23", "스프린트학교", "초등학교", 120], ["", "오류학교", "초등학교", "not-number"]],
        )
        importer = SchoolImport(source)

        self.assertEqual(importer.mapping["school_name"], "학교 이름")
        invalid = failed_rows(importer)
        self.assertEqual(len(invalid), 1)
        destination = os.path.join(self.temporary.name, "failed.xlsx")
        export_failed_rows(invalid, destination)
        self.assertEqual(load_workbook(destination).active.max_row, 2)

        result = importer.run()
        self.assertEqual(result["imported"], 1)
        self.assertEqual(result["failed"], 1)

    def test_crm_import_reuses_sales_activity_service_validation(self):
        school = self.workbook("school.xlsx", ["schoolcode", "schoolname"], [["S23", "Sprint School"]])
        SchoolImport(school).run()
        crm = self.workbook(
            "crm.xlsx",
            ["학교코드", "활동일자", "활동 유형", "담당자", "메모", "후속일", "상태"],
            [["S23", "2026-07-22", "Visit", "Kim", "Meeting", "2026-07-30", "Lead"]],
        )

        result = CRMImport(crm).run()

        self.assertEqual(result["status"], "SUCCESS")
        self.assertEqual(result["imported"], 1)

    def test_extended_history_is_persistent_without_database_columns(self):
        history_path = os.path.join(self.temporary.name, "history.json")
        store = ImportRunStore(history_path)
        store.record(
            PROFILES["crm"], "crm.xlsx", 12,
            {"status": "PARTIAL", "imported": 10, "failed": 2, "elapsed": 1.25},
        )

        history = ImportRunStore(history_path).history()

        self.assertEqual(history[0]["rows"], 12)
        self.assertEqual(history[0]["success"], 10)
        self.assertEqual(history[0]["failed"], 2)
        self.assertEqual(history[0]["duration"], 1.25)


if __name__ == "__main__":
    unittest.main()
