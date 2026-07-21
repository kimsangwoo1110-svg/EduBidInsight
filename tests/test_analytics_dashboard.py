import os
import tempfile
import unittest
from contextlib import closing

from openpyxl import load_workbook

from gui.dashboard import dashboard_kpi_values, purchase_cycle_text
from services import database
from services.analytics_service import AnalyticsService
from services.contract_service import ContractService
from services.dashboard_export_service import DashboardExportService
from services.project_service import ProjectService
from services.rule_service import RuleService


class AnalyticsDashboardTest(unittest.TestCase):
    def setUp(self):
        self.temp_directory = tempfile.TemporaryDirectory()
        self.original_db_name = database.DB_NAME
        database.DB_NAME = os.path.join(self.temp_directory.name, "analytics.db")
        database.create_database()
        with closing(database.get_connection()) as connection, connection:
            connection.execute("DELETE FROM rules")
        self.school_code = "S-ANALYTICS"
        self.seed_data()

    def tearDown(self):
        database.DB_NAME = self.original_db_name
        self.temp_directory.cleanup()

    def seed_data(self):
        ProjectService.create(
            self.school_code,
            "교실 개선",
            "시설",
            "완료",
            100_000_000,
            2024,
            2024,
            "테스트",
        )
        ProjectService.create(
            self.school_code,
            "디지털 교육",
            "디지털",
            "예정",
            200_000_000,
            2026,
            2026,
            "테스트",
        )
        for values in (
            ("2024-01-01", "태블릿", "업체A", 2, 100_000),
            ("2024-04-10", "태블릿", "업체A", 1, 300_000),
            ("2025-04-10", "노트북", "업체B", 3, 600_000),
        ):
            ContractService.save(
                school_code=self.school_code,
                school_name="분석학교",
                contract_date=values[0],
                product=values[1],
                category="기자재",
                vendor=values[2],
                quantity=values[3],
                amount=values[4],
                source_file="analytics.csv",
            )
        RuleService.create(
            "대형 계약",
            {"field": "amount", "operator": "gte", "value": 500_000},
            "대형 계약 후속 제안",
            45,
            "50만원 이상 계약",
        )
        RuleService.create(
            "예정 프로젝트",
            {"field": "status", "operator": "equals", "value": "예정"},
            "사업 사전 제안",
            30,
            "예정 사업",
        )

    def test_school_summary_and_contract_statistics(self):
        summary = AnalyticsService.school_summary(self.school_code)

        self.assertEqual(summary["kpis"]["projects"], 2)
        self.assertEqual(summary["kpis"]["project_budget"], 300_000_000)
        self.assertEqual(summary["kpis"]["contracts"], 3)
        self.assertEqual(summary["kpis"]["contract_amount"], 1_000_000)
        self.assertEqual(summary["contract_summary"]["average_amount"], 1_000_000 / 3)
        self.assertEqual(summary["contract_summary"]["total_quantity"], 6)
        self.assertEqual(summary["contract_summary"]["vendor_count"], 2)

    def test_vendor_and_product_statistics(self):
        vendors = AnalyticsService.vendor_statistics(self.school_code)
        products = AnalyticsService.product_statistics(self.school_code)

        self.assertEqual(vendors[0]["vendor"], "업체B")
        self.assertEqual(vendors[0]["amount"], 600_000)
        self.assertEqual(vendors[0]["share"], 60)
        self.assertEqual(products[0]["product"], "노트북")
        self.assertEqual(products[1]["count"], 2)
        self.assertEqual(products[1]["quantity"], 3)

    def test_yearly_trend_and_purchase_cycle(self):
        trend = AnalyticsService.yearly_trend(self.school_code)
        cycle = AnalyticsService.purchase_cycle(self.school_code)

        by_year = {row["year"]: row for row in trend}
        self.assertEqual(by_year[2024]["project_count"], 1)
        self.assertEqual(by_year[2024]["contract_count"], 2)
        self.assertEqual(by_year[2025]["contract_amount"], 600_000)
        self.assertEqual(by_year[2026]["project_budget"], 200_000_000)
        self.assertEqual(cycle["average_days"], 232.5)
        self.assertEqual(cycle["median_days"], 232.5)
        self.assertEqual(cycle["last_purchase_date"], "2025-04-10")
        self.assertEqual(cycle["status"], "분석 완료")

    def test_opportunity_score_combines_projects_and_contracts(self):
        opportunity = AnalyticsService.opportunity_score(self.school_code)

        self.assertEqual(opportunity["score"], 45)
        self.assertEqual(opportunity["raw_total"], 75)
        self.assertEqual(opportunity["opportunity_count"], 2)
        self.assertEqual(opportunity["priority"], "보통")
        self.assertEqual(
            {insight["target_type"] for insight in opportunity["insights"]},
            {"project", "contract"},
        )

    def test_dashboard_ui_helpers(self):
        summary = AnalyticsService.school_summary(self.school_code)
        cards = dashboard_kpi_values(summary)

        self.assertEqual(len(cards), 6)
        self.assertEqual(cards[0], ("프로젝트", "2건"))
        self.assertEqual(cards[-1], ("Opportunity Score", "45/100"))
        self.assertIn("평균 232.5일", purchase_cycle_text(summary["purchase_cycle"]))

    def test_excel_dashboard_export(self):
        summary = AnalyticsService.school_summary(self.school_code)
        file_path = os.path.join(self.temp_directory.name, "dashboard.xlsx")

        DashboardExportService.export_excel(file_path, "분석학교", summary)
        workbook = load_workbook(file_path, read_only=True, data_only=True)
        try:
            self.assertEqual(
                workbook.sheetnames,
                ["요약", "품목 통계", "업체 통계", "연도별 추이", "영업 기회"],
            )
            self.assertEqual(workbook["요약"]["B1"].value, "분석학교")
            self.assertEqual(workbook["품목 통계"]["A2"].value, "노트북")
        finally:
            workbook.close()

    def test_pdf_dashboard_export(self):
        summary = AnalyticsService.school_summary(self.school_code)
        file_path = os.path.join(self.temp_directory.name, "dashboard.pdf")

        DashboardExportService.export_pdf(file_path, "분석학교", summary)
        with open(file_path, "rb") as pdf_file:
            content = pdf_file.read()

        self.assertTrue(content.startswith(b"%PDF-1.4"))
        self.assertIn(b"/UniKS-UCS2-H", content)
        self.assertTrue(content.rstrip().endswith(b"%%EOF"))
        self.assertGreater(len(content), 1000)

    def test_empty_school_has_safe_zero_values(self):
        summary = AnalyticsService.school_summary("EMPTY")

        self.assertEqual(summary["contract_summary"]["total_count"], 0)
        self.assertEqual(summary["purchase_cycle"]["status"], "데이터 없음")
        self.assertEqual(summary["opportunity"]["score"], 0)
        self.assertEqual(summary["opportunity"]["priority"], "없음")


if __name__ == "__main__":
    unittest.main()
