import os
import tempfile
import unittest
from contextlib import closing
from datetime import date

from openpyxl import load_workbook

from gui.dashboard import business_insight_values
from services import database
from services.contract_service import ContractService
from services.dashboard_export_service import DashboardExportService
from services.insight_service import InsightService
from services.project_service import ProjectService
from services.rule_service import RuleService


class InsightServiceTest(unittest.TestCase):
    def setUp(self):
        self.temp_directory = tempfile.TemporaryDirectory()
        self.original_db_name = database.DB_NAME
        database.DB_NAME = os.path.join(self.temp_directory.name, "insight.db")
        database.create_database()
        with closing(database.get_connection()) as connection, connection:
            connection.execute("DELETE FROM rules")
        self.school_code = "S-BI"
        self.seed_data()

    def tearDown(self):
        database.DB_NAME = self.original_db_name
        self.temp_directory.cleanup()

    def seed_data(self):
        ProjectService.create(
            self.school_code,
            "AI 디지털 교실",
            "AI 교육",
            "예정",
            100_000_000,
            2026,
            2026,
            "테스트",
        )
        for contract_date, amount in (("2025-01-01", 900_000), ("2025-04-01", 100_000)):
            ContractService.save(
                school_code=self.school_code,
                school_name="BI학교",
                contract_date=contract_date,
                product="태블릿",
                category="디지털",
                vendor="집중업체",
                quantity=1,
                amount=amount,
                source_file="bi.csv",
            )
        RuleService.create(
            "예정",
            {"field": "status", "operator": "equals", "value": "예정"},
            "사전 제안",
            40,
            "예정 프로젝트",
        )
        RuleService.create(
            "계약",
            {"field": "amount", "operator": "gte", "value": 500_000},
            "후속 제안",
            50,
            "대형 계약",
        )

    def test_recommendation_history_schema(self):
        with closing(database.get_connection()) as connection:
            columns = [
                row[1]
                for row in connection.execute("PRAGMA table_info(recommendation_history)")
            ]
        self.assertEqual(
            columns,
            [
                "id",
                "school_code",
                "created_at",
                "score",
                "summary",
                "recommended_products",
                "next_action",
                "risk",
            ],
        )

    def test_summarize_school_persists_decodable_history(self):
        insight = InsightService.summarize_school(self.school_code)
        history = InsightService.history(self.school_code)

        self.assertIsNotNone(insight["history_id"])
        self.assertEqual(history[0]["id"], insight["history_id"])
        self.assertEqual(history[0]["score"], 50)
        self.assertIn("프로젝트 1건", history[0]["summary"])
        self.assertIsInstance(history[0]["recommended_products"], list)
        self.assertIsInstance(history[0]["next_action"], list)
        self.assertIsInstance(history[0]["risk"], list)

    def test_recommendation_explanation_and_next_action(self):
        analytics = InsightService.summarize_school(
            self.school_code, persist=False
        )["analytics"]
        recommendations = InsightService.recommend_products(
            self.school_code, analytics
        )
        explanation = InsightService.explain_score(self.school_code, analytics)
        risks = InsightService.risk_analysis(
            self.school_code, analytics, today=date(2025, 5, 1)
        )
        actions = InsightService.next_sales_action(
            self.school_code, analytics, recommendations, risks
        )

        self.assertEqual(recommendations[0]["product"], "AI 교육 솔루션")
        self.assertTrue(any(item["product"] == "태블릿" for item in recommendations))
        self.assertEqual(explanation["score"], 50)
        self.assertIn("영업 기회 2건", explanation["text"])
        self.assertEqual(actions[0]["timing"], "7일 이내")
        self.assertIn("맞춤 제안서", actions[1]["action"])

    def test_risk_timeline_and_priority_matrix(self):
        insight = InsightService.summarize_school(self.school_code, persist=False)
        risks = InsightService.risk_analysis(
            self.school_code,
            insight["analytics"],
            today=date(2026, 1, 1),
        )

        self.assertTrue(any("구매 집중도" in risk["risk"] for risk in risks))
        self.assertTrue(any("예상 구매 시점" in risk["risk"] for risk in risks))
        self.assertEqual(
            {event["type"] for event in insight["opportunity_timeline"]},
            {"계약", "프로젝트", "예상"},
        )
        self.assertTrue(insight["priority_matrix"])
        self.assertIn(
            insight["priority_matrix"][0]["quadrant"],
            {"즉시 실행", "전략 추진", "빠른 대응", "관찰"},
        )

    def test_dashboard_business_insight_helper(self):
        insight = InsightService.summarize_school(self.school_code, persist=False)
        values = business_insight_values(insight)

        self.assertIn("Opportunity Score", values["summary"])
        self.assertIn("Opportunity Score 50점", values["explanation"])
        self.assertEqual(values["recommendations"], len(insight["recommended_products"]))
        self.assertGreaterEqual(values["actions"], 2)

    def test_business_insight_excel_export(self):
        insight = InsightService.summarize_school(self.school_code, persist=False)
        path = os.path.join(self.temp_directory.name, "bi.xlsx")

        DashboardExportService.export_excel(path, "BI학교", insight["analytics"])
        workbook = load_workbook(path, read_only=True, data_only=True)
        try:
            self.assertIn("BI 요약", workbook.sheetnames)
            self.assertIn("BI 추천", workbook.sheetnames)
            self.assertIn("BI 위험과 행동", workbook.sheetnames)
            self.assertIn("기회 타임라인", workbook.sheetnames)
            self.assertIn("우선순위 매트릭스", workbook.sheetnames)
            self.assertIn("프로젝트 1건", workbook["BI 요약"]["B1"].value)
        finally:
            workbook.close()

    def test_business_insight_pdf_export(self):
        insight = InsightService.summarize_school(self.school_code, persist=False)
        path = os.path.join(self.temp_directory.name, "bi.pdf")

        DashboardExportService.export_pdf(path, "BI학교", insight["analytics"])
        with open(path, "rb") as pdf_file:
            content = pdf_file.read()

        encoded_heading = "Business Insight".encode("utf-16-be").hex().upper().encode("ascii")
        self.assertIn(encoded_heading, content)
        self.assertIn(b"/UniKS-UCS2-H", content)
        self.assertTrue(content.rstrip().endswith(b"%%EOF"))


if __name__ == "__main__":
    unittest.main()
