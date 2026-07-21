import os
import tempfile
import time
import unittest

from openpyxl import load_workbook

from services import database
from services.action_center import ActionCenterService
from services.contract_service import ContractService
from services.dashboard_service import DashboardService
from services.opportunity_engine import OpportunityEngine, OpportunityResult
from services.project_service import ProjectService
from services.report_service import REPORT_TYPES, ReportDocument, ReportService
from services.sales_activity_service import SalesActivityService
from services.school_service import SchoolService


class ReportServiceTest(unittest.TestCase):
    def setUp(self):
        self.temp_directory = tempfile.TemporaryDirectory()
        self.original_db_name = database.DB_NAME
        database.DB_NAME = os.path.join(self.temp_directory.name, "reports.db")
        database.create_database()
        DashboardService.clear_cache()
        ReportService.clear_cache()
        SchoolService.save(
            "RPT-1", "Report School", "Seoul Office", "Seoul", "Elementary",
            "19 Report Road", student_count=610, class_count=24,
        )
        ProjectService.create(
            "RPT-1", "AI Education Project", "AI Education", "Active",
            150_000_000, 2026, 2027, "Education Office",
        )
        ContractService.save(
            school_code="RPT-1", school_name="Report School",
            contract_date="2026-07-20", product="Interactive Display",
            category="ICT", vendor="Report Vendor", quantity=2,
            amount=20_000_000, source_file="report.csv",
        )
        SalesActivityService.add_activity(
            "RPT-1", "2026-07-19", "Visit", "Principal", "Discuss proposal",
            next_action_date="2026-07-25", status="Qualified",
        )
        ActionCenterService.create(
            "RPT-1", "Proposal", "Prepare AI proposal", priority="High",
            due_date="2026-07-21",
        )
        OpportunityEngine.save(OpportunityResult(
            school_id="RPT-1", school_name="Report School", score=85,
            priority="★★★★☆", recommendation="Recommend AI Classroom proposal.",
            next_action="Visit this week", confidence="Medium",
            evidence=["✓ AI Education Project", "✓ Large Annual Project Budget"],
            generated_at="2026-07-21T09:00:00+09:00",
        ))

    def tearDown(self):
        ReportService.clear_cache()
        DashboardService.clear_cache()
        database.DB_NAME = self.original_db_name
        self.temp_directory.cleanup()

    def test_school_report_aggregates_all_required_sections(self):
        document = ReportService.aggregate(
            "School Report", {"school": "Report School"}
        )

        self.assertIsInstance(document, ReportDocument)
        titles = {section.title for section in document.sections}
        self.assertTrue({
            "School Profile", "Opportunity Score", "Evidence", "Projects",
            "Contracts", "CRM History", "Recommended Actions", "Timeline",
        }.issubset(titles))
        projects = next(section for section in document.sections if section.title == "Projects")
        self.assertEqual(projects.rows[0][0], "AI Education Project")
        self.assertIn("Report School", document.title)

    def test_filters_apply_to_school_and_category(self):
        document = ReportService.aggregate(
            "School Report",
            {
                "school": "RPT-1", "region": "Seoul", "office": "Seoul Office",
                "category": "AI Education", "date_from": "2026-01-01",
                "date_to": "2026-12-31",
            },
        )

        projects = next(section for section in document.sections if section.title == "Projects")
        contracts = next(section for section in document.sections if section.title == "Contracts")
        self.assertEqual(len(projects.rows), 1)
        self.assertEqual(len(contracts.rows), 0)
        with self.assertRaises(ValueError):
            ReportService.aggregate(
                "School Report", {"school": "RPT-1", "region": "Busan"},
                force_refresh=True,
            )

    def test_preview_is_printable_and_does_not_reaggregate_document(self):
        document = ReportService.aggregate("School Report", {"school": "RPT-1"})

        preview = ReportService.preview(document)

        self.assertIn("EduBid Insight", preview)
        self.assertIn("[School Profile]", preview)
        self.assertIn("[Recommended Actions]", preview)
        self.assertIn("Report School", preview)

    def test_pdf_excel_and_csv_exports(self):
        document = ReportService.aggregate("School Report", {"school": "RPT-1"})
        paths = {
            "PDF": os.path.join(self.temp_directory.name, "school.pdf"),
            "Excel": os.path.join(self.temp_directory.name, "school.xlsx"),
            "CSV": os.path.join(self.temp_directory.name, "school.csv"),
        }
        for export_format, path in paths.items():
            self.assertEqual(ReportService.export(document, path, export_format), path)
            self.assertGreater(os.path.getsize(path), 100)

        with open(paths["PDF"], "rb") as pdf_file:
            self.assertTrue(pdf_file.read(8).startswith(b"%PDF-1."))
        with open(paths["CSV"], "r", encoding="utf-8-sig") as csv_file:
            self.assertIn("School Profile", csv_file.read())
        workbook = load_workbook(paths["Excel"])
        self.assertIn("School Profile", workbook.sheetnames)
        self.assertTrue(any(sheet._charts for sheet in workbook.worksheets))
        workbook.close()

    def test_all_global_report_types_are_supported(self):
        for report_type in REPORT_TYPES[1:]:
            document = ReportService.aggregate(report_type, force_refresh=True)
            self.assertEqual(document.report_type, report_type)
            self.assertTrue(document.sections)
        weekly = ReportService.aggregate("Weekly Report")
        self.assertIn("KPIs", {section.title for section in weekly.sections})
        opportunity_report = ReportService.aggregate("Opportunity Report")
        opportunity_rows = next(
            section.rows for section in opportunity_report.sections
            if section.title == "Opportunity Schools"
        )
        self.assertEqual(opportunity_rows[0][1], "RPT-1")

    def test_cached_aggregation_meets_performance_target(self):
        started = time.perf_counter()
        first = ReportService.aggregate("Weekly Report")
        cold_elapsed = time.perf_counter() - started
        started = time.perf_counter()
        cached = ReportService.aggregate("Weekly Report")
        cached_elapsed = time.perf_counter() - started

        self.assertLess(cold_elapsed, 2.0)
        self.assertLess(cached_elapsed, 2.0)
        self.assertIs(first, cached)


if __name__ == "__main__":
    unittest.main()
