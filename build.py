"""One-command production build for EduBid Insight Personal v1.0."""

import argparse
import ctypes
import html
import os
import shutil
import subprocess
import sys
import time
from pathlib import Path
from tkinter import TclError


ROOT = Path(__file__).resolve().parent
BUILD_DIR = ROOT / "build"
DIST_DIR = ROOT / "dist"
RELEASE_DIR = ROOT / "release"
OUTPUT_PDF_DIR = ROOT / "output" / "pdf"
EXE_NAME = "EduBidInsight.exe"
VERSION = "1.0.0"


def log(message):
    print(f"[build] {message}", flush=True)


def run(command, *, env=None, timeout=None):
    log(" ".join(str(part) for part in command))
    subprocess.run(
        [str(part) for part in command],
        cwd=ROOT,
        env=env,
        timeout=timeout,
        check=True,
    )


def require_build_dependencies():
    missing = []
    for module in ("PyInstaller", "PIL", "reportlab", "pypdf", "pdfplumber"):
        try:
            __import__(module)
        except ImportError:
            missing.append(module)
    if missing:
        raise RuntimeError(
            "Missing build dependencies: "
            + ", ".join(missing)
            + ". Run: python -m pip install -r requirements-build.txt"
        )


def generate_icon():
    """Generate the established EB mark as PNG and multi-resolution ICO."""
    from PIL import Image, ImageDraw, ImageFont

    assets = ROOT / "assets"
    assets.mkdir(parents=True, exist_ok=True)
    canvas = Image.new("RGBA", (1024, 1024), (0, 0, 0, 0))
    draw = ImageDraw.Draw(canvas)
    draw.rounded_rectangle((80, 80, 944, 944), radius=205, fill="#4F78A6")
    draw.rounded_rectangle((116, 116, 908, 908), radius=175, outline="#6F9ACA", width=14)
    font_path = Path(os.environ.get("WINDIR", r"C:\Windows")) / "Fonts" / "seguisb.ttf"
    font = ImageFont.truetype(str(font_path), 350) if font_path.is_file() else ImageFont.load_default()
    text = "EB"
    box = draw.textbbox((0, 0), text, font=font)
    x = (1024 - (box[2] - box[0])) / 2 - box[0]
    y = (1024 - (box[3] - box[1])) / 2 - box[1] - 8
    draw.text((x, y), text, font=font, fill="white")
    png_path = assets / "edubid_icon.png"
    ico_path = assets / "edubid.ico"
    canvas.save(png_path, optimize=True)
    canvas.save(
        ico_path,
        format="ICO",
        sizes=[(16, 16), (24, 24), (32, 32), (48, 48), (64, 64), (128, 128), (256, 256)],
    )
    log(f"Generated icon resources: {png_path.name}, {ico_path.name}")


def generate_import_templates():
    """Create valid, styled starter workbooks for every bundled importer."""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    from services.import_center import PROFILES

    definitions = tuple(
        (profile.template_name, profile.fields, profile.labels)
        for profile in PROFILES.values()
    )
    template_directory = ROOT / "templates"
    template_directory.mkdir(parents=True, exist_ok=True)
    for filename, columns, labels in definitions:
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "Import Template"
        worksheet.append([labels.get(column, column) for column in columns])
        worksheet.freeze_panes = "A2"
        worksheet.auto_filter.ref = f"A1:{get_column_letter(len(columns))}1"
        for index, column in enumerate(columns, start=1):
            cell = worksheet.cell(1, index)
            cell.font = Font(name="Malgun Gothic", size=10, bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="4F78A6")
            cell.alignment = Alignment(horizontal="center", vertical="center")
            worksheet.column_dimensions[get_column_letter(index)].width = max(
                14, len(labels.get(column, column)) * 2 + 4
            )
        worksheet.row_dimensions[1].height = 24
        workbook.save(template_directory / filename)
        workbook.close()
    log(f"Generated {len(definitions)} official bundled import templates")


def _guide_story(markdown_text):
    from reportlab.lib import colors
    from reportlab.lib.enums import TA_LEFT
    from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
    from reportlab.lib.units import mm
    from reportlab.platypus import PageBreak, Paragraph, Preformatted, Spacer

    styles = getSampleStyleSheet()
    regular = "MalgunGothic"
    bold = "MalgunGothicBold"
    body = ParagraphStyle(
        "GuideBody", parent=styles["BodyText"], fontName=regular,
        fontSize=10.5, leading=16, textColor=colors.HexColor("#253142"),
        spaceAfter=5 * mm, alignment=TA_LEFT,
    )
    title = ParagraphStyle(
        "GuideTitle", parent=body, fontName=bold, fontSize=24, leading=30,
        textColor=colors.HexColor("#345F8D"), spaceAfter=10 * mm,
    )
    heading = ParagraphStyle(
        "GuideHeading", parent=body, fontName=bold, fontSize=15, leading=20,
        textColor=colors.HexColor("#345F8D"), spaceBefore=5 * mm, spaceAfter=3 * mm,
        keepWithNext=True,
    )
    bullet = ParagraphStyle(
        "GuideBullet", parent=body, leftIndent=7 * mm, firstLineIndent=-4 * mm,
        bulletIndent=1 * mm, spaceAfter=2 * mm,
    )
    code = ParagraphStyle(
        "GuideCode", parent=body, fontName="Courier", fontSize=8.5, leading=12,
        backColor=colors.HexColor("#F0F2F5"), borderPadding=7,
        spaceBefore=2 * mm, spaceAfter=4 * mm,
    )

    story = []
    code_lines = []
    in_code = False

    def clean_inline(value):
        return html.escape(value.replace("**", "").replace("`", "").replace("\\", "/"))

    for raw_line in markdown_text.splitlines():
        line = raw_line.rstrip()
        if line.startswith("```"):
            if in_code:
                story.append(Preformatted("\n".join(code_lines), code))
                code_lines = []
            in_code = not in_code
            continue
        if in_code:
            code_lines.append(line)
        elif line.startswith("# "):
            story.append(Paragraph(clean_inline(line[2:]), title))
            story.append(Paragraph("Education Sales Intelligence Platform", body))
        elif line.startswith("## "):
            heading_text = line[3:]
            if heading_text == "CRM Action Center":
                story.append(PageBreak())
            story.append(Paragraph(clean_inline(heading_text), heading))
        elif line.startswith("- "):
            story.append(Paragraph("• " + clean_inline(line[2:]), bullet))
        elif line:
            story.append(Paragraph(clean_inline(line), body))
        else:
            story.append(Spacer(1, 1.5 * mm))
    return story


def generate_user_guide_pdf():
    """Generate and verify the release PDF from the maintained Markdown guide."""
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.units import mm
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont
    from reportlab.platypus import SimpleDocTemplate
    from pypdf import PdfReader

    windows_fonts = Path(os.environ.get("WINDIR", r"C:\Windows")) / "Fonts"
    regular = windows_fonts / "malgun.ttf"
    bold = windows_fonts / "malgunbd.ttf"
    if not regular.is_file() or not bold.is_file():
        raise RuntimeError("Malgun Gothic fonts are required to generate the Korean user guide")
    pdfmetrics.registerFont(TTFont("MalgunGothic", str(regular)))
    pdfmetrics.registerFont(TTFont("MalgunGothicBold", str(bold)))

    OUTPUT_PDF_DIR.mkdir(parents=True, exist_ok=True)
    destination = OUTPUT_PDF_DIR / "USER_GUIDE.pdf"

    def page_chrome(canvas, document):
        canvas.saveState()
        canvas.setStrokeColor(colors.HexColor("#DDE3EA"))
        canvas.line(18 * mm, 16 * mm, 192 * mm, 16 * mm)
        canvas.setFont("MalgunGothic", 8)
        canvas.setFillColor(colors.HexColor("#667085"))
        canvas.drawString(18 * mm, 10 * mm, f"EduBid Insight Personal v{VERSION}")
        canvas.drawRightString(192 * mm, 10 * mm, f"Page {document.page}")
        canvas.restoreState()

    document = SimpleDocTemplate(
        str(destination), pagesize=A4, rightMargin=18 * mm, leftMargin=18 * mm,
        topMargin=18 * mm, bottomMargin=22 * mm,
        title="EduBid Insight Personal v1.0 User Guide",
        author="Kim Sangwoo",
        subject="Education Sales Intelligence Platform",
    )
    source = (ROOT / "USER_GUIDE.md").read_text(encoding="utf-8")
    document.build(_guide_story(source), onFirstPage=page_chrome, onLaterPages=page_chrome)
    reader = PdfReader(str(destination))
    if not reader.pages or destination.stat().st_size < 5_000:
        raise RuntimeError("Generated USER_GUIDE.pdf failed validation")
    log(f"Generated USER_GUIDE.pdf ({len(reader.pages)} pages)")
    return destination


def prepare_tcl_runtime():
    """Prepare compatible Tcl/Tk scripts for the Python 3.14 Windows build."""
    source_root = Path(sys.base_prefix) / "tcl"
    runtime_root = BUILD_DIR / "tcl-runtime"
    if runtime_root.exists():
        shutil.rmtree(runtime_root)
    runtime_root.mkdir(parents=True)
    for directory in ("tcl8.6", "tk8.6"):
        source = source_root / directory
        if not source.is_dir():
            raise RuntimeError(f"Tcl/Tk runtime is missing: {source}")
        shutil.copytree(source, runtime_root / directory)

    replacements = (
        (runtime_root / "tcl8.6" / "init.tcl", "package require -exact Tcl 8.6.15", "package require Tcl 8.6"),
        (runtime_root / "tk8.6" / "tk.tcl", "package require -exact Tk  8.6.15", "package require Tk 8.6"),
    )
    for path, original, replacement in replacements:
        text = path.read_text(encoding="utf-8")
        if original not in text:
            raise RuntimeError(f"Unexpected Tcl/Tk runtime version declaration: {path}")
        path.write_text(text.replace(original, replacement, 1), encoding="utf-8")
    environment = os.environ.copy()
    environment["TCL_LIBRARY"] = str(runtime_root / "tcl8.6")
    environment["TK_LIBRARY"] = str(runtime_root / "tk8.6")
    log("Prepared validated Tcl/Tk runtime for packaging")
    return environment


def build_executable():
    environment = prepare_tcl_runtime()
    run(
        [sys.executable, "-m", "PyInstaller", "--noconfirm", "--clean", "EduBidInsight.spec"],
        env=environment,
    )
    executable = DIST_DIR / EXE_NAME
    if not executable.is_file() or executable.stat().st_size < 1_000_000:
        raise RuntimeError(f"PyInstaller output is missing or unexpectedly small: {executable}")
    return executable


def _safe_recreate(directory):
    resolved = directory.resolve()
    if resolved.parent != ROOT:
        raise RuntimeError(f"Refusing to replace directory outside project root: {resolved}")
    if resolved.exists():
        shutil.rmtree(resolved)
    resolved.mkdir(parents=True)


def assemble_release(executable, guide_pdf):
    _safe_recreate(RELEASE_DIR)
    shutil.copy2(executable, RELEASE_DIR / EXE_NAME)
    shutil.copy2(ROOT / "packaging" / "README.txt", RELEASE_DIR / "README.txt")
    shutil.copy2(ROOT / "CHANGELOG.md", RELEASE_DIR / "CHANGELOG.md")
    shutil.copy2(guide_pdf, RELEASE_DIR / "USER_GUIDE.pdf")
    shutil.copy2(ROOT / "LICENSE", RELEASE_DIR / "LICENSE")
    for name in ("config", "data", "backups"):
        (RELEASE_DIR / name).mkdir()
    shutil.copy2(
        ROOT / "packaging" / "default_config" / "app_settings.json",
        RELEASE_DIR / "config" / "app_settings.json",
    )
    log(f"Assembled release folder: {RELEASE_DIR}")


def _application_windows():
    from ctypes import wintypes

    found = []
    enum_callback = ctypes.WINFUNCTYPE(ctypes.c_bool, wintypes.HWND, wintypes.LPARAM)

    def visit(handle, _parameter):
        owner_pid = wintypes.DWORD()
        ctypes.windll.user32.GetWindowThreadProcessId(handle, ctypes.byref(owner_pid))
        if ctypes.windll.user32.IsWindowVisible(handle):
            length = ctypes.windll.user32.GetWindowTextLengthW(handle)
            title = ctypes.create_unicode_buffer(length + 1)
            ctypes.windll.user32.GetWindowTextW(handle, title, length + 1)
            if "EduBid Insight" in title.value:
                found.append((handle, owner_pid.value, title.value))
        return True

    ctypes.windll.user32.EnumWindows(enum_callback(visit), 0)
    return found


def verify_windows_metadata(executable):
    import pefile

    pe = pefile.PE(str(executable))
    try:
        if pe.OPTIONAL_HEADER.Subsystem != 2:
            raise RuntimeError("Executable is not using the Windows GUI subsystem")
        strings = {}
        for file_info in getattr(pe, "FileInfo", []):
            for block in file_info:
                if getattr(block, "Key", b"") == b"StringFileInfo":
                    for table in block.StringTable:
                        strings.update({key.decode(): value.decode() for key, value in table.entries.items()})
        expected = {
            "ProductName": "EduBid Insight",
            "CompanyName": "Kim Sangwoo",
            "FileDescription": "Education Sales Intelligence Platform",
            "ProductVersion": VERSION,
        }
        for key, value in expected.items():
            if strings.get(key) != value:
                raise RuntimeError(f"Windows metadata mismatch for {key}: {strings.get(key)!r}")
    finally:
        pe.close()
    log("Verified GUI subsystem and Windows version properties")


def verify_reports(database_path, output_directory):
    from services import database
    from services.migration_service import MigrationService
    from services.report_service import ReportService

    database.configure_database(str(database_path))
    MigrationService(str(database_path)).migrate()
    output_directory.mkdir(parents=True, exist_ok=True)
    document = ReportService.aggregate("Weekly Report", force_refresh=True)
    targets = {
        "PDF": output_directory / "weekly_report.pdf",
        "Excel": output_directory / "weekly_report.xlsx",
        "CSV": output_directory / "weekly_report.csv",
    }
    for format_name, path in targets.items():
        ReportService.export(document, str(path), format_name)
        if not path.is_file() or path.stat().st_size == 0:
            raise RuntimeError(f"Report smoke test failed: {path}")
    log("Verified PDF, Excel, and CSV report generation")


def smoke_test(executable):
    profile = BUILD_DIR / "smoke-profile"
    if profile.exists():
        shutil.rmtree(profile)
    profile.mkdir(parents=True)
    environment = os.environ.copy()
    environment["LOCALAPPDATA"] = str(profile)
    environment["APPDATA"] = str(profile)
    environment["EDUBID_PORTABLE"] = "0"
    existing_windows = {item[0] for item in _application_windows()}
    process = subprocess.Popen([str(executable)], cwd=RELEASE_DIR, env=environment)
    app_root = profile / "EduBidInsight"
    database_path = app_root / "data" / "edubid.db"
    settings_path = app_root / "config" / "app_settings.json"
    handle = None
    deadline = time.monotonic() + 35
    try:
        while time.monotonic() < deadline:
            if process.poll() is not None:
                raise RuntimeError(f"Executable exited during startup with code {process.returncode}")
            candidates = [item for item in _application_windows() if item[0] not in existing_windows]
            handle = candidates[0][0] if candidates else None
            if handle and database_path.is_file() and settings_path.is_file():
                break
            time.sleep(0.25)
        else:
            raise RuntimeError("Executable startup/dashboard/settings/database smoke test timed out")
        time.sleep(2)
        ctypes.windll.user32.PostMessageW(handle, 0x0010, 0, 0)
        process.wait(timeout=20)
    finally:
        if process.poll() is None:
            process.terminate()
            process.wait(timeout=10)
    backups = list((app_root / "backups").glob("edubid_*_exit.db"))
    if process.returncode != 0 or not backups:
        raise RuntimeError("Graceful executable exit or automatic backup verification failed")
    verify_reports(database_path, BUILD_DIR / "smoke-reports")
    log("Verified executable startup, dashboard, settings, database creation, and exit backup")


def main():
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--skip-tests", action="store_true", help="Skip regression tests")
    parser.add_argument("--skip-smoke", action="store_true", help="Skip executable smoke tests")
    args = parser.parse_args()

    require_build_dependencies()
    generate_icon()
    generate_import_templates()
    guide_pdf = generate_user_guide_pdf()
    if not args.skip_tests:
        run([sys.executable, "-m", "pytest", "-q"])
    executable = build_executable()
    assemble_release(executable, guide_pdf)
    verify_windows_metadata(RELEASE_DIR / EXE_NAME)
    if not args.skip_smoke:
        smoke_test(RELEASE_DIR / EXE_NAME)
    log(f"SUCCESS - EduBid Insight Personal {VERSION}")
    log(f"Release: {RELEASE_DIR}")
    return 0


if __name__ == "__main__":
    try:
        raise SystemExit(main())
    except (RuntimeError, subprocess.CalledProcessError, subprocess.TimeoutExpired, TclError) as error:
        print(f"[build] FAILED: {error}", file=sys.stderr)
        raise SystemExit(1)
