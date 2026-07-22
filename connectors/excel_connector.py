"""Generic Excel/CSV transport connector with injectable persistence."""

from __future__ import annotations

import csv
import io
from pathlib import Path

from openpyxl import load_workbook

from connectors.base_connector import BaseConnector, ConnectorMetadata


class ExcelConnector(BaseConnector):
    metadata = ConnectorMetadata(
        key="excel", name="Excel / CSV", profile_key="crm",
        description="Official workbook and CSV import transport", is_mock=False,
    )

    def __init__(self, file_path=None, sheet_name=None, import_handler=None):
        super().__init__()
        self.file_path = Path(file_path).resolve() if file_path else None
        self.sheet_name = sheet_name
        self.import_handler = import_handler
        self._workbook = None
        self._headers = []

    def connect(self):
        if self.file_path is None or not self.file_path.is_file():
            raise ValueError("an existing Excel or CSV file is required")
        extension = self.file_path.suffix.lower()
        if extension not in {".xlsx", ".csv"}:
            raise ValueError("only .xlsx and .csv files are supported")
        if extension == ".xlsx":
            self._workbook = load_workbook(self.file_path, read_only=True, data_only=True)
            selected = self.sheet_name or self._workbook.sheetnames[0]
            if selected not in self._workbook.sheetnames:
                self._workbook.close(); self._workbook = None
                raise ValueError(f"unknown worksheet: {selected}")
            self.sheet_name = selected
        else:
            self.sheet_name = "CSV"
        self.connected = True
        return self

    def disconnect(self):
        if self._workbook is not None:
            self._workbook.close()
            self._workbook = None
        self.connected = False

    def validate(self):
        if not self.connected:
            return False
        rows = self._source_rows()
        try:
            self._headers = [str(value or "").strip() for value in next(rows, ())]
        finally:
            close = getattr(rows, "close", None)
            if close:
                close()
        if not any(self._headers):
            raise ValueError("the workbook header row is empty")
        return True

    def fetch(self):
        if not self.connected:
            raise RuntimeError("connector is not connected")
        rows = self._source_rows()
        try:
            headers = [str(value or "").strip() for value in next(rows, ())]
            return [
                {header: values[index] if index < len(values) else None for index, header in enumerate(headers) if header}
                for values in rows
                if any(value not in (None, "") for value in values)
            ]
        finally:
            close = getattr(rows, "close", None)
            if close:
                close()

    def import_data(self, records):
        if self.import_handler is None:
            return {"status": "READY", "imported": 0, "skipped": len(records), "failed": 0}
        result = self.import_handler([dict(record) for record in records])
        if not isinstance(result, dict):
            raise TypeError("Excel import_handler must return a result mapping")
        return result

    def _source_rows(self):
        if self.file_path.suffix.lower() == ".xlsx":
            yield from self._workbook[self.sheet_name].iter_rows(values_only=True)
            return
        content = None
        for encoding in ("utf-8-sig", "cp949"):
            try:
                content = self.file_path.read_text(encoding=encoding)
                break
            except UnicodeDecodeError:
                continue
        if content is None:
            raise ValueError("CSV encoding must be UTF-8 or CP949")
        yield from csv.reader(io.StringIO(content))
